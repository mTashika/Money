import os
import openpyxl
import CONST as C
from Tools import get_previous_month,write_list_to_excel,protect_ws,unprotect_cell
from tkinter import messagebox

class ManageExcelFile:
    def __init__(self,folder_path,year,month):
        self.file_name = f'{year}.xlsx'
        self.sheet_name = month
        
        self.folder_path = folder_path
        self.file_path = os.path.join(self.folder_path, self.file_name)
        
        self.wb_valid = False
        self.ws_valid = False
        
        self.get_workbook_and_sheet()
        self.save_wb()
        
    def get_wb_ws(self):
        if not self.wb_valid:
            print('Workbook not valid')
            return None,None
        if not self.ws_valid:
            print('Worksheet not valid')
            return None,None
        return self.workbook,self.ws
    
    def get_workbook_and_sheet(self):
        file_exists, sheet_exists = self.check_excel_file_and_sheet()
        if not file_exists:
            print('Workbook creation..')
            self.init_excel_year()
            print('Workbook creation - OK')
            self.wb_valid = True if self.workbook else False
        else:
            print('Workbook Loading..')
            self.workbook = openpyxl.load_workbook(self.file_path)
            print('Workbook Loading - OK')
            self.wb_valid = True if self.workbook else False
        if not sheet_exists:
            self.ws = self.workbook.create_sheet(title=self.sheet_name)
            self.ws_valid = True if self.ws else False
        else:
            if messagebox.askyesno("Confirmation", f"This sheet ({self.sheet_name}) already exist in {self.file_name}, do you want to delete it all to recreate it ?"):
                self.workbook.remove(self.workbook[self.sheet_name])
                self.ws = self.workbook.create_sheet(title=self.sheet_name)
                self.ws_valid = True if self.ws else False
            else:
                self.ws_valid = False
        
    def save_wb(self):
        if self.wb_valid and self.ws_valid:
            try:
                self.workbook.save(self.file_path)
                print('Workbook saved')
            except PermissionError:
                messagebox.showerror('Error','Permission denied\nTry to close the excel first')
        else:
            print('Workbook not saved')
                
            
    def check_excel_file_and_sheet(self):
        """
        Vérifie si un fichier Excel existe, est accessible, et si un onglet spécifique existe dans ce fichier.
        :param self.file_path: Chemin du fichier Excel.
        :param self.sheet_name: Nom de l'onglet à vérifier.
        :return: Tuple (file_exists, sheet_exists) :
                - file_exists: True si le fichier existe et est accessible, False sinon.
                - sheet_exists: True si l'onglet existe et est accessible, False sinon.
        """
        try:
            file_exists = os.path.isfile(self.file_path) and os.access(self.file_path, os.R_OK)
            sheet_exists = False
            if file_exists:
                try:
                    workbook = openpyxl.load_workbook(self.file_path, read_only=True)
                    if self.sheet_name in workbook.sheetnames:
                        sheet_exists = True
                except Exception as e:
                    print(f"Une erreur s'est produite lors de l'accès au fichier ou à l'onglet: {e}")
            return file_exists, sheet_exists
        except PermissionError:
            raise PermissionError(f"Permission refusée: Impossible d'accéder au fichier '{self.path_excel}'. Veuillez vérifier que le fichier n'est pas ouvert dans une autre application.")
            
        except FileNotFoundError:
            raise FileNotFoundError(f"Fichier non trouvé: Impossible de localiser le chemin spécifié '{self.path_excel}'.")
            
        except Exception as e:
            raise Exception(f"Une erreur s'est produite: {e}")
        
    def init_excel_year(self):
        '''
        Create and save the excel file and add the Tool sheet
        '''
        # create excel file
        self.workbook = openpyxl.Workbook()
        # create Tool sheet
        self.init_sheet = self.workbook.active
        self.init_sheet.title = C.INIT_SHEET_NAME
        
        # fill init sheet
        self.init_sheet['A1'].style="Accent1"
        self.init_sheet[C.INIT_SHEET_MONTH_CELL[1]].style="Accent1"
        self.init_sheet[C.INIT_SHEET_SOLD_CELL[1]].style="Input"
        self.init_sheet.merge_cells('A2:B2')

        self.init_sheet['A1'] = f'{C.INIT_SHEET_SOLD_TITLE}'
        self.init_sheet[C.INIT_SHEET_MONTH_CELL[1]] = get_previous_month(self.sheet_name)
        self.init_sheet[f'{C.INIT_SHEET_CAT1_COL}1'] = f'{C.INIT_SHEET_CAT1_TITLE}'
        self.init_sheet[f'{C.INIT_SHEET_CAT2_COL}1'] = f'{C.INIT_SHEET_CAT2_TITLE}'

        write_list_to_excel(self.init_sheet,C.CAT_1,[2,C.INIT_SHEET_CAT1_COL])
        write_list_to_excel(self.init_sheet,C.CAT_2,[2,C.INIT_SHEET_CAT2_COL])

        self.init_sheet.column_dimensions['A'].width = 22
        self.init_sheet.column_dimensions[C.INIT_SHEET_MONTH_CELL[0]].width = 15
        self.init_sheet.column_dimensions[C.INIT_SHEET_CAT1_COL].width = 15
        self.init_sheet.column_dimensions[C.INIT_SHEET_CAT2_COL].width = 15

        self.init_sheet['A1'].alignment = C.ALIGN_CENTER
        self.init_sheet[C.INIT_SHEET_MONTH_CELL[1]].alignment = C.ALIGN_CENTER
        self.init_sheet[f'{C.INIT_SHEET_CAT1_COL}1'].alignment = C.ALIGN_CENTER
        self.init_sheet[f'{C.INIT_SHEET_CAT2_COL}1'].alignment = C.ALIGN_CENTER
        self.init_sheet[C.INIT_SHEET_SOLD_CELL[1]].alignment = C.ALIGN_CENTER

        protect_ws(self.init_sheet)
        unprotect_cell(self.init_sheet,C.INIT_SHEET_SOLD_CELL[1])
        

        