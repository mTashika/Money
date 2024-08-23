from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Créer un nouveau classeur et sélectionner la feuille active
wb = Workbook()
ws = wb.active

# Définir les valeurs valides dans la plage B1:B10
valeurs_valides = ["Option1", "Option2", "Option3", "Option4", "Option5", "Option6", "Option7", "Option8", "Option9", "Option10"]
for i, valeur in enumerate(valeurs_valides, start=1):
    ws[f"B{i}"] = valeur

# Créer une validation de données de type liste avec la plage B1:B10
valeurs_directes = "Option1,Option2,Option3,Option4,Option5,Option6,Option7,Option8,Option9,Option10"
validation_list = DataValidation(
    type="list",
    formula1=f'"{valeurs_directes}"',  # Liste directement dans la formule
    showDropDown=True,
    showErrorMessage=True
)
validation_list.error = "La valeur doit être l'une des options disponibles dans B1:B10."
validation_list.errorTitle = "Valeur non valide"

# Appliquer la validation à la cellule A1
ws.add_data_validation(validation_list)
validation_list.add(ws["A1"])

# Enregistrer le classeur
wb.save("exemple_menu_deroulant.xlsx")
