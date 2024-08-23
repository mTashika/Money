import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, NamedStyle,Alignment,Font,PatternFill
import random
import psutil
import tkinter as tk
from tkinter import messagebox
from openpyxl.utils.cell import range_boundaries
import customtkinter as ctk
from tkinter import filedialog

import CONST as C

# global values
COL_TO_RESET = [6,7,8,9,10,11]
S = None

# some functions
def verif_nan(val):
    return ~(isinstance(val, (int,float)) and np.isnan(val))
def round_p(val,n=2):
    return round(val,n)
def round_e(val,n=1):
    return round(val,n)

# Style
class StyleCell():
    def __init__(self,wb):
        self.wb = wb
        self.thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        self.style_cat1_list = []
        colors_cat1 = ["FFC3C0", "FFD1B8", "FFE6B8", "E6FFB8", "C0FFC3", "B8FFD1", "B8E6FF", "B8C3FF", "D1B8FF", "FFB8E6"]
        
        
        style_title_name = "style_title"
        self.style_title = NamedStyle(name=style_title_name,
                        font=Font(color="FF000000"),
                        fill=PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid"),
                        border=self.thin_border,
                        alignment=Alignment(horizontal='center', vertical='center')
                        ) if style_title_name not in wb.named_styles else style_title_name
        
        style_title_val = "style_title_val"
        self.style_title_val = NamedStyle(name=style_title_val,
                        font=Font(color="FF000000"),
                        fill=PatternFill(start_color="FFB3E5FC", end_color="FFB3E5FC", fill_type="solid"),
                        border=self.thin_border,
                        alignment=Alignment(horizontal='center', vertical='center'),
                        number_format='0 €'
                        ) if style_title_val not in wb.named_styles else style_title_val
        
        style_cat1 = "style_cat1"
        self.style_cat1 = NamedStyle(name=style_cat1,
                        font=Font(color="FF000000"),
                        fill=PatternFill(start_color=colors_cat1[0], end_color=colors_cat1[0], fill_type="solid"),
                        border=self.thin_border,
                        alignment=Alignment(horizontal='center', vertical='center'),
                        ) if style_cat1 not in wb.named_styles else style_cat1
        
        for i,c in enumerate(colors_cat1):
            self.gen_style_cat1(f'style_cat{i+1}',c)
        
    def gen_style_cat1(self,name,color):
        if name not in self.wb.named_styles:
            self.style_cat1_list.append(
                                    NamedStyle(name=name,
                                            font=Font(color="FF000000"),
                                            fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                                            border=self.thin_border,
                                            alignment=Alignment(horizontal='center', vertical='center'),
                                            )
                                    )
        else : 
            self.style_cat1_list.append(name)

class Sheet():
    def __init__(self,path,sheet = None):
        self.file_path = path
        self.sheet_name = sheet
        self.data = None
        self.cell_sold_init = [1,0]
        self.c_months = 0
        self.col_cat1 = self.c_months
        self.col_cat2 = self.col_cat1+1
        self.c_note = self.col_cat2+1
        self.c_val = self.c_note+1
        self.c_w1 = self.c_val+3
        self.c_w2 = self.c_w1+1
        self.c_w3 = self.c_w2+1
        self.c_w4 = self.c_w3+1
        self.c_w5 = self.c_w4+1
        
        self.months_obj = []
        
        self.read_excel()
        self.verif_bal_init()
        
    def is_file_open(self,file_path):
        for proc in psutil.process_iter():
            try:
                files = proc.open_files()
                for item in files:
                    if file_path == item.path:
                        return True
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
        return False

    def read_excel(self):
        try:
            if self.sheet_name:
                self.data = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None)
            else:
                self.data = pd.read_excel(self.file_path, header=None)
        except PermissionError as e :
            print("Error: The Excel file is already open by another application.")
            self.popup_alrdy_open()
            raise e
        except Exception as e:
            print("Error reading Excel file:", e)
            raise e
            
    def get_all_months(self):
        self.dic_m = {}
        for row, val in self.data.iterrows():
            v = val[self.c_months]
            if isinstance(v,str):
                if C.BALISE_NEW_MONTH in v:
                    m_txt = v.replace(C.BALISE_NEW_MONTH, '')
                    if m_txt not in self.dic_m:
                        self.dic_m[m_txt] = row
            
    def verif_bal_init(self):
        v = pd.read_excel(self.file_path, sheet_name=C.INIT_SHEET_NAME, header=None)
        val = v.iloc[self.cell_sold_init[0],self.cell_sold_init[1]]
        if not isinstance(val,(int,float)) or not verif_nan(val):
            raise Exception ('Balance value not at the expected spot')
        
    def get_bal_prev(self):
        if self.months_obj == []:
            v = pd.read_excel(self.file_path, sheet_name=C.INIT_SHEET_NAME, header=None)
            val = v.iloc[self.cell_sold_init[0],self.cell_sold_init[1]]
            return val
        return self.months_obj[-1].bal_usr
        
    def gen_data(self):
        self.get_all_months()         
        def get_row_end_month(c_name,dic):
            dic= list(dic.items())
            for i, (key, _) in enumerate(dic):
                if key == c_name:
                    if i + 1 >= len(dic):
                        return None
                    else:
                        return dic[i + 1][1]
            
        for month, row_s in self.dic_m.items():
            row_e = get_row_end_month(month,self.dic_m)
            m_data = self.data.iloc[row_s:row_e,:]
            
            self.months_obj.append(Month(m_data,self.get_bal_prev()))

    def clear_zone(self):        
        
        mcr_coord_list = [mcr.coord for mcr in self.sh.merged_cells.ranges]
    
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            if min_col in COL_TO_RESET:
                self.sh.unmerge_cells(mcr)
                
        
        for idx, col in enumerate(self.sh.iter_cols(), start=1):
            if idx in COL_TO_RESET:
                for cell in col:
                    cell.value = None
                    cell.style = 'Normal'
        
    def write(self):
        global S
        self.wb = load_workbook(self.file_path)
        self.sh = self.wb[self.sheet_name]
        self.clear_zone()
        
        S = StyleCell(self.wb)
        
        for month in self.months_obj:
            row = month.data.index.to_list()[0]+1
            self.sh.cell(row,self.c_w1, month.txt_m_tot)
            self.sh.cell(row,self.c_w1).style = S.style_title
            self.sh.cell(row,self.c_w2, month.txt_m_exp_tot)
            self.sh.cell(row,self.c_w2).style = S.style_title
            self.sh.cell(row,self.c_w3, month.txt_m_inc_tot)
            self.sh.cell(row,self.c_w3).style = S.style_title
            
            self.sh.cell(row,self.c_w4, month.txt_m_tot_real)
            self.sh.cell(row,self.c_w4).style = S.style_title
            self.sh.cell(row,self.c_w4).border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thin'))
            self.sh.cell(row,self.c_w5, month.txt_m_exp_real)
            self.sh.cell(row,self.c_w5).style = S.style_title
            self.sh.cell(row,self.c_w5).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thin'))
            self.sh.cell(row,self.c_w5+1, month.txt_m_inc_real)
            self.sh.cell(row,self.c_w5+1).style = S.style_title
            self.sh.cell(row,self.c_w5+1).border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thin'))
            row +=1
            self.sh.cell(row,self.c_w1, round_e(month.tot))
            self.sh.cell(row,self.c_w1).style = S.style_title_val
            self.sh.cell(row,self.c_w2, round_e(month.tot_exp))
            self.sh.cell(row,self.c_w2).style = S.style_title_val
            self.sh.cell(row,self.c_w3, round_e(month.tot_inc))
            self.sh.cell(row,self.c_w3).style = S.style_title_val
            
            self.sh.cell(row,self.c_w4, round_e(month.tot_real))
            self.sh.cell(row,self.c_w4).style = S.style_title_val
            self.sh.cell(row,self.c_w4).border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thick'))
            self.sh.cell(row,self.c_w5, round_e(month.tot_exp_real))
            self.sh.cell(row,self.c_w5).style = S.style_title_val
            self.sh.cell(row,self.c_w5).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thick'))
            self.sh.cell(row,self.c_w5+1, round_e(month.tot_inc_real))
            self.sh.cell(row,self.c_w5+1).style = S.style_title_val
            self.sh.cell(row,self.c_w5+1).border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thin'),bottom=Side(style='thick'))
            
            self.sh.cell(month.row_bal+2,self.col_cat1+1, month.txt_m_bal_real)
            self.sh.cell(month.row_bal+2,self.col_cat1+1).style = 'Output'
            self.sh.cell(month.row_bal+2,self.col_cat2+1, month.bal_pred)
            self.sh.cell(month.row_bal+2,self.col_cat2+1).style = 'Output'
            
            self.sh.cell(month.row_bal+1,self.c_note+1, month.txt_m_verif)
            self.sh.cell(month.row_bal+1,self.c_val+1, month.verif)
            
            if month.verif:
                self.sh.cell(month.row_bal+1,self.c_note+1).style = 'Good'
                self.sh.cell(month.row_bal+1,self.c_val+1).style = 'Good'
            else : 
                self.sh.cell(month.row_bal+1,self.c_note+1).style = 'Bad'
                self.sh.cell(month.row_bal+1,self.c_val+1).style = 'Bad'
            
            self.sh.merge_cells(start_row=month.row_bal+1, start_column=self.c_note+1, end_row=month.row_bal+2, end_column=self.c_note+1)
            self.sh.merge_cells(start_row=month.row_bal+1, start_column=self.c_note+2, end_row=month.row_bal+2, end_column=self.c_note+2)
            
            self.sh.cell(month.row_bal+1,self.c_note+1).alignment=Alignment(horizontal='center', vertical='center')
            self.sh.cell(month.row_bal+1,self.c_val+1).alignment=Alignment(horizontal='center', vertical='center')
            
            month.write_cat1(self.sh,month.data.index.to_list()[0]+4,self.c_w1)
            
        self.wb.save(self.file_path)
    
    def popup_alrdy_open(self):
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)  # Mettre la fenêtre au-dessus de toutes les autres
        messagebox.showwarning("Fichier Excel ouvert", "Le fichier Excel est déjà ouvert par une autre application.", parent=root)
        root.destroy()

class Month():
    def __init__(self,data,bal_prev):
        self.data=data
        self.name = data.iloc[0,0].replace(C.BALISE_NEW_MONTH, '')
        self.cat1=[]
        self.tot = 0
        self.tot_exp = 0
        self.tot_inc = 0
        self.tot_exp_real = 0
        self.tot_inc_real = 0
        self.tot_real = 0
        self.bal_pred = 0
        self.bal_usr = 0
        self.bal_prev = bal_prev
        
        # txt
        self.txt_m_tot = 'Total transaction'
        self.txt_m_exp_tot = 'Dépense totale'
        self.txt_m_inc_tot = 'Revenue totale'
        self.txt_m_tot_real= 'Total réelle'
        self.txt_m_exp_real= 'Dépense réelle'
        self.txt_m_inc_real = 'Revenue réel'
        self.txt_m_bal_real= 'Solde estimé :'
        self.txt_m_verif='Verif'
        
        self.row_s = None
        self.find_start()
        
        self.d_cat1 = self.data.loc[self.row_s:,0]
        self.d_val = self.data.loc[self.row_s:,3]
        
        
        self.update_tot()
        self.update_bal()
        self.create_cat1()

    def find_start(self):
        for row,val in self.data.iterrows():
            if isinstance(val[0],str):
                if not C.BALISE_NEW_MONTH in val[0]:
                    self.row_s = row
                    return
                
    def update_tot(self):
        for _,val in self.data.iterrows():
            euro = val[3]
            ok_real = True if val[0]not in C.NOT_REAL_ACTION else False
            if isinstance(euro,(int,float)) and verif_nan(euro):
                self.tot+=euro
                if ok_real:
                    self.tot_real+=euro
                if euro > 0:
                    self.tot_inc+=euro
                    if ok_real:
                        self.tot_inc_real+=euro
                    
                else:
                    self.tot_exp+=euro
                    if ok_real:
                        self.tot_exp_real+=euro 
    
    def update_bal(self):
        for row,val in self.data.iterrows():
            if isinstance(val[0],str):
                if C.BALISE_SOLD_MONTH in val[0]:
                    self.bal_usr = val[1]
                    self.row_bal = row
                    break
        self.bal_pred = self.bal_prev + self.tot
        self.verif_bal()
        
    def verif_bal(self):
        marg_p = 10 # %
        marg_q = 2
        verif_p =  np.abs(self.bal_usr*(1-marg_p/100)) <= np.abs(self.bal_pred) <= np.abs(self.bal_usr*(1+marg_p/100))
        verif_q = self.bal_pred <= self.bal_usr + marg_q and self.bal_pred >= self.bal_usr - marg_q 
        if verif_q or verif_p:
            self.verif = True
        else:
            self.verif = False
            
    def create_cat1(self):
        # get all categories with the dataframe corresponding to it
        dic_cat1 = {}
        for id,val in self.d_cat1.items():
            if verif_nan(val):
                if C.BALISE_SOLD_MONTH in val:
                    break
                if val not in dic_cat1:
                    dic_cat1[val] = self.data.loc[id]
                else:
                    dic_cat1[val]=pd.concat([dic_cat1[val], self.data.loc[id]], axis=1)
                    
        # create categories 
        for _,df in dic_cat1.items():
            self.cat1.append(Cat(df.transpose(),self.tot_exp_real,self.tot_inc_real))
    
    def sort_cat1(self):
            
        def custom_sort(cat):
            if cat.tot_euro >= 0:
                return (1, cat.tot_euro) # 1 pour dire de le mettre au dessus de la list, tot_euro pour que sorted fasse un rangement avec aussi
            else:
                return (0, cat.tot_euro)
        self.cat1 = sorted(self.cat1, key=custom_sort)
            
    
    def write_cat1(self,sh,row_s,col_s):
        self.sort_cat1()
        for cat in self.cat1:
            cat.write(sh,row_s,col_s)
            if len(cat.data.shape) == 1:
                add_row = 1
            else:
                add_row = len(cat.cat2)+1
            row_s+=add_row

class Cat():
    def __init__(self,data,exp_tot,inc_tot):
        self.inc_tot = inc_tot
        self.exp_tot = exp_tot
        self.data = data
        self.txt_total = 'Total'
        self.disp_p = True
        self.tot_p = None
        self.cat2=[]
        
        if len(self.data.shape) == 1:
            self.name = self.data[0]
            self.d_cat2 = self.data.loc[1]
            self.d_euro = self.data.loc[3]
            self.tot_euro = self.d_euro
            
            if self.name in C.NOT_REAL_ACTION:
                self.disp_p = False
            if self.disp_p:
                if self.tot_euro>0:
                    ref = self.inc_tot
                else:
                    ref = self.exp_tot
                self.tot_p = self.tot_euro*100/ref
                
            if verif_nan(self.d_cat2):
                self.txt_total = self.d_cat2
            return
        
        self.name = self.data.iloc[0,0]
        if self.name in C.NOT_REAL_ACTION:
            self.disp_p = False
        self.d_cat2 = self.data.loc[:,1]
        self.d_euro = self.data.loc[:,3]
        self.tot_euro = 0
        self.tot_p = None
        self.get_tot()
        self.create_cat2()
        
        
    def get_tot(self):
        for id,val in self.d_euro.items():
            self.tot_euro+=val
        
        if self.disp_p:
            if self.tot_euro>0:
                ref = self.inc_tot
            else:
                ref = self.exp_tot
            self.tot_p = self.tot_euro*100/ref
            
    def create_cat2(self):
        # check if multiple cat2
        if self.d_cat2.nunique(dropna=False) == 1:  # only 1 sub categorie
            if verif_nan(self.d_cat2.iloc[0]):
                self.txt_total = self.d_cat2.iloc[0]
        else:
            # get all categories with the dataframe corresponding to it
            dic_cat2 = {}
            for id,val in self.d_cat2.items():
                if verif_nan(val):
                    if val not in dic_cat2:
                        dic_cat2[val] = self.data.loc[id]
                    else:
                        dic_cat2[val]=pd.concat([dic_cat2[val], self.data.loc[id]], axis=1)
                        
            # create categories 
            for _,df in dic_cat2.items():
                self.cat2.append(Cat2(df.transpose(),self.tot_euro,self.disp_p))
        
    def write(self,sh,row_s,col_s):
        r_s = row_s
        sh.cell(row_s,col_s,self.name)
        sh.cell(row_s,col_s).font = Font(bold=True)
        col_s+=1
        
        for cat2 in self.cat2:
            cat2.write(sh,row_s,col_s)
            row_s+=1
        cat_st = random.choice(S.style_cat1_list)
        sh.cell(r_s,col_s-1).style =cat_st
        sh.merge_cells(start_row=r_s, start_column=col_s-1, end_row=row_s, end_column=col_s-1)
        
        sh.cell(row_s,col_s,self.txt_total)
        sh.cell(row_s,col_s).style =cat_st
        
        sh.cell(row_s,col_s+1,round_e(self.tot_euro))
        sh.cell(row_s,col_s+1).style =cat_st
        sh.cell(row_s,col_s+1).number_format='0 €'
        
        if self.disp_p:
            sh.cell(row_s,col_s+2,round_p(self.tot_p))
            sh.cell(row_s,col_s+2).style =cat_st
            sh.cell(row_s,col_s+2).number_format='0 \%'

class Cat2():
    def __init__(self,data,tot_cat,disp_p):
        self.data = data
        self.tot_cat = tot_cat
        self.tot_p = None
        self.disp_p = disp_p
        
        if len(self.data.shape) == 1:
            self.name = self.data[1]
            self.d_euro = self.data.loc[3]
            self.tot_euro = self.d_euro
            if self.disp_p:
                self.tot_p = self.tot_euro*100/self.tot_cat
            return
        
        self.name = self.data.iloc[0,1]
        self.d_euro = self.data.loc[:,3]
        self.tot_euro = 0
        self.tot_p = 0
        
        self.get_tot()
        
    def get_tot(self):
        for id,val in self.d_euro.items():
            self.tot_euro+=val
        if self.disp_p:
            self.tot_p = self.tot_euro*100/self.tot_cat
    
    
    def write(self,sh,row_s,col_s):
        sh.cell(row_s,col_s, self.name)
        sh.cell(row_s,col_s).style = 'Calculation'
        sh.cell(row_s,col_s+1, round_e(self.tot_euro))
        sh.cell(row_s,col_s+1).style = 'Calculation'
        sh.cell(row_s,col_s+1).number_format='0 €'
        sh.cell(row_s,col_s+1).alignment=Alignment(horizontal='center', vertical='center')
                
        if self.disp_p:  
            sh.cell(row_s,col_s+2, round_p(self.tot_p))
            sh.cell(row_s,col_s+2).style = 'Calculation'
            sh.cell(row_s,col_s+2).number_format='0 \%'
            sh.cell(row_s,col_s+2).alignment=Alignment(horizontal='center', vertical='center')


def compute_code_month(excel_path,sheet_name):
    print('\n*** Start ***')
    # try: 
    sheet = Sheet(excel_path,sheet_name)
    print("Sheet opened")
    sheet.gen_data()
    print("Data generated")
    sheet.write()
    print("Success !")
    # except Exception as e:
    #     print(e)
    # finally:
    #     print('*** End ***\n')




