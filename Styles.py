from openpyxl.styles import Side, Border, NamedStyle,Alignment,Font,PatternFill

class StyleCell():
    def __init__(self,wb):
        self.wb = wb
        self.thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        self.style_cat1_list = []
        colors_cat1 = ["FFC3C0", "FFD1B8", "FFE6B8", "E6FFB8", "C0FFC3", "B8FFD1", "B8E6FF", "B8C3FF", "D1B8FF", "FFB8E6"]
        
        
        style_title_name = "style_title"
        self.style_title = NamedStyle(name=style_title_name,
                        font=Font(color="FF000000"),
                        fill=PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid"),
                        border=self.thin_border,
                        alignment=Alignment(horizontal='center', vertical='center')
                        ) if style_title_name not in wb.named_styles else style_title_name
        
        style_title_val = "style_title_val"
        self.style_title_val = NamedStyle(name=style_title_val,
                        font=Font(color="FF000000"),
                        fill=PatternFill(start_color="FFB3E5FC", end_color="FFB3E5FC", fill_type="solid"),
                        border=self.thin_border,
                        alignment=Alignment(horizontal='center', vertical='center'),
                        number_format='0 €'
                        ) if style_title_val not in wb.named_styles else style_title_val
        
        style_cat1 = "style_cat1"
        self.style_cat1 = NamedStyle(name=style_cat1,
                        font=Font(color="FF000000"),
                        fill=PatternFill(start_color="95b3d7", end_color="95b3d7", fill_type="solid"),
                        border=self.thin_border,
                        alignment=Alignment(horizontal='left', vertical='center'),
                        ) if style_cat1 not in wb.named_styles else style_cat1
        style_cat2 = "style_cat2"
        self.style_cat2 = NamedStyle(name=style_cat2,
                        font=Font(color="FF000000"),
                        fill=PatternFill(start_color="da9694", end_color="da9694", fill_type="solid"),
                        border=self.thin_border,
                        alignment=Alignment(horizontal='left', vertical='center'),
                        ) if style_cat2 not in wb.named_styles else style_cat2
        
        for i,c in enumerate(colors_cat1):
            self.gen_style_cat1(f'style_cat{i+1}',c)
            
        

        
        
        
        
        
        
        
        
        
        
    def gen_style_cat1(self,name,color):
        if name not in self.wb.named_styles:
            self.style_cat1_list.append(
                                    NamedStyle(name=name,
                                            font=Font(color="FF000000"),
                                            fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                                            border=self.thin_border,
                                            alignment=Alignment(horizontal='center', vertical='center'),
                                            )
                                    )
        else : 
            self.style_cat1_list.append(name)