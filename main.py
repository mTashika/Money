import customtkinter as ctk
import os
from tkinter import filedialog
from tkinter import messagebox
from Tools import center_window,update_init_sold,is_file_closed
from openpyxl import load_workbook

from Extract_from_pdf import PDFExtraction
from Manage_excel_file import ManageExcelFile
from Write_fo_in_excel import WriteFo
from Styles import StyleCell
from Compute_month import compute_code_month
# Initialiser l'application CustomTkinter
ctk.set_appearance_mode("System")  # Modes: "System" (default), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Thèmes: "blue" (default), "dark-blue", "green"

GREY = "#999999"
WHITE = "white"
class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Configurer la fenêtre
        self.title("Gestion des Chemins")

        # Centrer la fenêtre sur l'écran
        center_window(self,800, 600)

        # Variables pour stocker les chemins
        self.pdf_path = ctk.StringVar(value="D:/One Drive/OneDrive/Bureau/relevejuin.pdf")
        self.excel_dir_path = ctk.StringVar(value="D:/.Perso_local/Projects/Project_Money")
        self.year_var = ctk.StringVar(value="2022")
        self.year_var.trace_add("write", self.update_combo_box)
        self.month_var = ctk.StringVar(value="juin")
        self.checkbox_create_var = ctk.BooleanVar()
        self.checkbox_update_var = ctk.BooleanVar()

        # Relier la fermeture par la croix à la fonction annuler
        self.protocol("WM_DELETE_WINDOW", self.cancel)
        
        frame_excel = ctk.CTkFrame(self)
        frame_choice = ctk.CTkFrame(self)
        frame_excel.pack()
        frame_choice.pack()
        
        frame_choice_create = ctk.CTkFrame(frame_choice)
        frame_choice_update = ctk.CTkFrame(frame_choice)
        frame_choice_create.pack(side='left')
        frame_choice_update.pack(side='left')
        
        # frame excel
        self.label_excel_dir = ctk.CTkLabel(frame_excel, text="Chemin du dossier Excel:")
        self.label_excel_dir.pack(pady=(20, 0), padx=20, anchor='w')
        self.entry_excel_dir = ctk.CTkEntry(frame_excel, textvariable=self.excel_dir_path, width=400)
        self.entry_excel_dir.pack(pady=10, padx=20, anchor='w')
        self.button_browse_excel_dir = ctk.CTkButton(frame_excel, text="Browse", command=self.browse_excel_dir)
        self.button_browse_excel_dir.pack(pady=10, padx=20, anchor='w')
        
        # frame_choice_create
        self.label_create_title = ctk.CTkLabel(frame_choice_create, text="Create")
        self.label_create_title.pack(pady=(20, 0), padx=20, anchor='w')
        
        self.label_pdf = ctk.CTkLabel(frame_choice_create, text="Chemin du fichier PDF:")
        self.label_pdf.pack(pady=(20, 0), padx=20, anchor='w')
        self.entry_pdf = ctk.CTkEntry(frame_choice_create, textvariable=self.pdf_path, width=400)
        self.entry_pdf.pack(pady=10, padx=20, anchor='w')
        self.button_browse_pdf = ctk.CTkButton(frame_choice_create, text="Browse", command=self.browse_pdf)
        self.button_browse_pdf.pack(pady=10, padx=20, anchor='w')
        
        self.checkbox_create = ctk.CTkCheckBox(frame_choice_create, text='', variable=self.checkbox_create_var, onvalue=True, offvalue=False,command=lambda: self.on_checkbox_toggle(self.checkbox_create_var))
        self.checkbox_create.pack(pady=10, padx=20, anchor='w')
        
        # frame_choice_update
        self.label_update_title = ctk.CTkLabel(frame_choice_update, text="Update")
        self.label_update_title.pack(pady=(20, 0), padx=20, anchor='w')
        
        self.label_year = ctk.CTkLabel(frame_choice_update, text="Year")
        self.label_year.pack(pady=(20, 0), padx=20, anchor='w')
        self.entry_year = ctk.CTkEntry(frame_choice_update, textvariable=self.year_var, width=400)
        self.entry_year.pack(pady=10, padx=20, anchor='w')
        self.label_month = ctk.CTkLabel(frame_choice_update, text="Month")
        self.label_month.pack(pady=(20, 0), padx=20, anchor='w')
        
        self.entry_month = ctk.CTkComboBox(frame_choice_update, variable=self.month_var, width=400,values='')
        self.entry_month.pack(pady=10, padx=20, anchor='w')
        
        self.checkbox_update = ctk.CTkCheckBox(frame_choice_update, text='', variable=self.checkbox_update_var, onvalue=True, offvalue=False,command=lambda: self.on_checkbox_toggle(self.checkbox_update_var))
        self.checkbox_update.pack(pady=10, padx=20, anchor='w')
        
        self.button_free = ctk.CTkButton(self, text="Free", command=self.free_year_wb)
        self.button_free.pack(side="right", padx=20, pady=20)
        
        # Boutons OK et Annuler
        self.button_ok = ctk.CTkButton(self, text="OK", command=self.ok)
        self.button_ok.pack(side="right", padx=20, pady=20)

        self.button_cancel = ctk.CTkButton(self, text="Annuler", command=self.cancel)
        self.button_cancel.pack(side="right", padx=10, pady=20)
        
    def browse_pdf(self):
        file_path = filedialog.askopenfilename(title="Sélectionner un fichier PDF", filetypes=[("PDF Files", "*.pdf")])
        if file_path:
            self.pdf_path.set(file_path)
    
    def browse_excel_dir(self):
        dir_path = filedialog.askdirectory(title="Sélectionner un dossier")
        if dir_path:
            self.excel_dir_path.set(dir_path)
            
    def get_wb_from_year_var(self):
        year = self.year_var.get()
        excel_file_path = os.path.join(self.excel_dir_path.get(), f'{year}.xlsx')
        return load_workbook(filename=excel_file_path)
        
    def get_existing_sheet_from_year(self):
        try:
            wb = self.get_wb_from_year_var()
            return wb.sheetnames
        except Exception:
            return ['']
    def free_year_wb(self):
        try:
            wb = self.get_wb_from_year_var()
            wb.close()
            messagebox.showinfo('Info','Workbook closed')
        except Exception:
            pass
    def update_combo_box(self, *args):
        sheet_names = self.get_existing_sheet_from_year()
        self.entry_month.configure(values=sheet_names)
        

    def disable_update(self):
        self.label_update_title.configure(text_color=GREY)
        self.label_year.configure(text_color=GREY)
        self.entry_year.configure(state="disabled",text_color=GREY)
        self.label_month.configure(text_color=GREY)
        self.entry_month.configure(state="disabled",text_color=GREY)
        self.checkbox_update.configure(state="normal")
        self.checkbox_update_var.set(False)
    def enable_update(self):
        self.label_update_title.configure(text_color=WHITE)
        self.label_year.configure(state="normal",text_color=WHITE)
        self.entry_year.configure(state="normal",text_color=WHITE)
        self.label_month.configure(state="normal",text_color=WHITE)
        self.entry_month.configure(state="normal",text_color=WHITE)
        self.checkbox_update.configure(state="normal")
        self.checkbox_update_var.set(True)
        
    def disable_create(self):
        self.label_create_title.configure(text_color=GREY)
        self.label_pdf.configure(text_color=GREY)
        self.entry_pdf.configure(state="disabled",text_color=GREY)
        self.button_browse_pdf.configure(state="disabled")
        self.checkbox_create.configure(state="normal")
        self.checkbox_create_var.set(False)
    def enable_create(self):
        self.label_create_title.configure(text_color=WHITE)
        self.label_pdf.configure(state="normal",text_color=WHITE)
        self.entry_pdf.configure(state="normal",text_color=WHITE)
        self.button_browse_pdf.configure(state="normal")
        self.checkbox_create.configure(state="normal")
        self.checkbox_create_var.set(True)
    
    def glob_enable_create(self):
        self.enable_create()
        self.disable_update()
    def glob_enable_update(self):
        self.enable_update()
        self.disable_create()
    
    def on_checkbox_toggle(self, checkbox_var):
    # Check if the event was triggered by the "Create" checkbox
        if checkbox_var == self.checkbox_create_var:
                self.glob_enable_create()
        # Check if the event was triggered by the "Update" checkbox
        elif checkbox_var == self.checkbox_update_var:
                self.glob_enable_update()


    # Fonction appelée lorsqu'on clique sur "Annuler" ou ferme la fenêtre
    def cancel(self):
        self.destroy()

    def ok(self):
        if self.checkbox_update_var.get() == False and self.checkbox_create_var.get() == True: # Create
            print('PDF extraction ..')
            self.extract_pdf_data()
            print('PDF extraction - OK')
            self.manage_excel_file()
            if self.wb and self.ws:
                print('Write in sheet ..')
                self.write_in_sheet()
                update_init_sold(self.wb,self.ws.title)
                print('Write in sheet - OK')
                
                self.wb.save(self.excel_file_path)
                print('Workbook saved - OK')
                messagebox.showinfo("Success", "Operation completed successfully!")
            
        elif self.checkbox_update_var.get() == True and self.checkbox_create_var.get() == False: # Update
            self.year = self.year_var.get()
            self.month = self.month_var.get().lower()
            self.excel_file_path = os.path.join(self.excel_dir_path.get(), f'{self.year}.xlsx')
            compute_code_month(self.excel_file_path,self.month)
        else:
            print('ERROR clique ok')
            

    def extract_pdf_data(self):
        extraction = PDFExtraction(self.pdf_path.get())
        self.fo = extraction.financial_operations
        self.month = extraction.month
        self.year = extraction.year
        
    def write_in_sheet(self):
        S = StyleCell(self.wb)
        WriteFo(self.fo,self.ws,S)
        self.ws['A1'] = self.pdf_path.get()
        
    def manage_excel_file(self):
        man = ManageExcelFile(self.excel_dir_path.get(),self.year,self.month)
        self.wb,self.ws = man.get_wb_ws()
        self.excel_file_path = man.file_path


if __name__ == "__main__":
    app = App()
    app.mainloop()
