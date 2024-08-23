import os
import openpyxl
import CONST as C
from Tools import get_previous_month,write_list_to_excel
from Styles import StyleCell as S
from Extract_from_pdf import PDFExtraction


class WriteFo:
    def __init__(self,fo,path_excel,month):
        self.file_path = path_excel
        self.sheet_name = month
        self.fo = fo
        self.month_start_row = C.MONTH_ROW_TITLE+C.START_MONTH[0]
        self.month_end_row = self.month_start_row + len(self.fo)+C.NB_MORE_FO_MONTH
        
        self.get_workbook_and_sheet()
        self.styles = S(self.workbook)
        self.init_excel_month()
        self.write_fo()
        self.workbook.save(self.file_path)
        
    def get_workbook_and_sheet(self):
        file_exists, sheet_exists = self.check_excel_file_and_sheet()
        if not file_exists:
            self.init_excel_year()
        else:
            self.workbook = openpyxl.load_workbook(self.file_path)
            
        if not sheet_exists:
            self.ws = self.workbook.create_sheet(title=self.sheet_name)
        else:
            print('? Sure to change this month ?')
            self.ws = self.workbook[self.sheet_name]
        
    def check_excel_file_and_sheet(self):
        """
        Vérifie si un fichier Excel existe, est accessible, et si un onglet spécifique existe dans ce fichier.
        :param self.file_path: Chemin du fichier Excel.
        :param self.sheet_name: Nom de l'onglet à vérifier.
        :return: Tuple (file_exists, sheet_exists) :
                - file_exists: True si le fichier existe et est accessible, False sinon.
                - sheet_exists: True si l'onglet existe et est accessible, False sinon.
        """
        try:
            file_exists = os.path.isfile(self.file_path) and os.access(self.file_path, os.R_OK)
            sheet_exists = False
            if file_exists:
                try:
                    workbook = openpyxl.load_workbook(self.file_path, read_only=True)
                    if self.sheet_name in workbook.sheetnames:
                        sheet_exists = True
                except Exception as e:
                    print(f"Une erreur s'est produite lors de l'accès au fichier ou à l'onglet: {e}")
            return file_exists, sheet_exists
        except PermissionError:
            raise PermissionError(f"Permission refusée: Impossible d'accéder au fichier '{self.path_excel}'. Veuillez vérifier que le fichier n'est pas ouvert dans une autre application.")
            
        except FileNotFoundError:
            raise FileNotFoundError(f"Fichier non trouvé: Impossible de localiser le chemin spécifié '{self.path_excel}'.")
            
        except Exception as e:
            raise Exception(f"Une erreur s'est produite: {e}")
        
    def init_excel_year(self):
        '''
        Create and save the excel file and add the Tool sheet
        '''
        # create excel file
        self.workbook = openpyxl.Workbook()
        # create Tool sheet
        self.init_sheet = self.workbook.active
        self.init_sheet.title = C.INIT_SHEET_NAME
        # fullfil init sheet
        self.init_sheet['A1'].style="Accent1"
        self.init_sheet['A2'].style="Accent1"

        self.init_sheet['A1'] = f'{C.INIT_SHEET_SOLD_TITLE} ({get_previous_month("May")})'
        self.init_sheet['B1'] = f'{C.INIT_SHEET_CAT1_TITLE}'
        self.init_sheet['C1'] = f'{C.INIT_SHEET_CAT2_TITLE}'

        write_list_to_excel(self.init_sheet,C.CAT_1,[2,2])
        write_list_to_excel(self.init_sheet,C.CAT_2,[2,3])

        self.init_sheet.column_dimensions['A'].width = 30
        self.init_sheet.column_dimensions['B'].width = 15
        self.init_sheet.column_dimensions['C'].width = 15

        self.init_sheet['A1'].alignment = C.ALIGN_CENTER
        self.init_sheet['B1'].alignment = C.ALIGN_CENTER
        self.init_sheet['C1'].alignment = C.ALIGN_CENTER
        
        
    def init_excel_month(self):
        self.init_cell_size_month()
        self.init_color_month()
        self.init_alignement_month()
        # self.init_boder_month()
        self.init_value_month()
        
    def init_cell_size_month(self):
        # column width
        self.ws.column_dimensions['A'].width = 34
        self.ws.column_dimensions['B'].width = 16
        self.ws.column_dimensions['C'].width = 26
        self.ws.column_dimensions['D'].width = 11
        self.ws.column_dimensions['E'].width = 2
        self.ws.column_dimensions['F'].width = 20
        self.ws.column_dimensions['G'].width = 20
        self.ws.column_dimensions['H'].width = 20
        self.ws.column_dimensions['I'].width = 20
        self.ws.column_dimensions['J'].width = 15
        self.ws.column_dimensions['K'].width = 15
        
        # merge
        self.ws.merge_cells(start_row=C.START_MONTH[0], start_column=C.START_MONTH[1],
                            end_row=C.START_MONTH[0]-1+C.MONTH_ROW_TITLE, end_column=C.START_MONTH[1]+3)
        
        self.ws.merge_cells(start_row=self.month_end_row , start_column=C.START_MONTH[1]+2,
                            end_row=self.month_end_row+1, end_column=C.START_MONTH[1]+2)
        self.ws.merge_cells(start_row=self.month_end_row, start_column=C.START_MONTH[1]+3,
                            end_row=self.month_end_row+1, end_column=C.START_MONTH[1]+3)
    def init_color_month(self):
        self.ws.cell(row=C.START_MONTH[0],column=C.START_MONTH[1]).style='Input'
        
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]).style='Normal'
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]+1).style='Note'
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]+2).style='Neutral'
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]+3).style='Neutral'
        
        self.ws.cell(row=self.month_end_row+1,column=C.START_MONTH[1]).style='Output'
        self.ws.cell(row=self.month_end_row+1,column=C.START_MONTH[1]+1).style='Output'
        for row in range(self.month_start_row,self.month_end_row):
            self.ws.cell(row=row,column=1).style = self.styles.style_cat1
            self.ws.cell(row=row,column=2).style = self.styles.style_cat2
            self.ws.cell(row=row,column=3).style = "Normal"
            self.ws.cell(row=row,column=4).style = "Note"
    def init_alignement_month(self):
        self.ws.cell(row=C.START_MONTH[0],column=C.START_MONTH[1]).alignment=C.ALIGN_CENTER
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]+2).alignment=C.ALIGN_CENTER
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]+3).alignment=C.ALIGN_CENTER
    # def init_boder_month(self):
    def init_value_month(self):
        self.ws['A1'] = self.file_path # HERE
        self.ws.cell(row=C.START_MONTH[0],column=C.START_MONTH[1]).value=f'{C.BALISE_NEW_MONTH}{self.sheet_name}'
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]).value = C.MONTH_SOLD_BALISE
        self.ws.cell(row=self.month_end_row,column=C.START_MONTH[1]+2).value = C.MONTH_VERIF
        self.ws.cell(row=self.month_end_row+1,column=C.START_MONTH[1]).value = C.MONTH_SOLD_EST

    def write_fo(self):
        for row,fo in enumerate(self.fo,start=self.month_start_row):
            self.ws.cell(row=row,column=C.START_MONTH[1]+2).value=fo.name
            self.ws.cell(row=row,column=C.START_MONTH[1]+3).value=fo.value


        
if __name__ == '__main__':
    path_pdf = r"D:\One Drive\OneDrive\Bureau\releve_CCP2142397R038_20240628.pdf"
    fo = PDFExtraction(path_pdf).financial_operations
    path_excel = "output.xlsx"
    WriteFo(fo,path_excel,"August")