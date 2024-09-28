from openpyxl.styles import Side, Border, NamedStyle, Alignment, Font, PatternFill

class StyleCell():
    def __init__(self, wb):
        self.wb = wb
        self.thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        self.style_cat1_list = []
        colors_cat1 = ["FFD1B8", "FFE6B8", "E6FFB8", "C0FFC3", "B8FFD1", "B8E6FF", "B8C3FF", "D1B8FF", "FFB8E6"]

        # Ajouter le style "style_title"
        style_title_name = "style_title"
        if style_title_name not in wb.named_styles:
            style_title = NamedStyle(
                name=style_title_name,
                font=Font(color="FF000000"),
                fill=PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid"),
                border=self.thin_border,
                alignment=Alignment(horizontal='center', vertical='center')
            )
            wb.add_named_style(style_title)

        # Ajouter le style "style_title_val"
        style_title_val = "style_title_val"
        if style_title_val not in wb.named_styles:
            style_title_val = NamedStyle(
                name=style_title_val,
                font=Font(color="FF000000"),
                fill=PatternFill(start_color="FFB3E5FC", end_color="FFB3E5FC", fill_type="solid"),
                border=self.thin_border,
                alignment=Alignment(horizontal='center', vertical='center'),
                number_format='0 €'
            )
            wb.add_named_style(style_title_val)

        # Ajouter le style "style_cat1"
        style_cat1 = "style_cat1"
        if style_cat1 not in wb.named_styles:
            style_cat1 = NamedStyle(
                name=style_cat1,
                font=Font(color="FF000000"),
                fill=PatternFill(start_color="FFC3C0", end_color="FFC3C0", fill_type="solid"),
                border=self.thin_border,
                alignment=Alignment(horizontal='center', vertical='center'),
            )
            wb.add_named_style(style_cat1)

        # Générer et ajouter des styles pour les catégories
        for i, c in enumerate(colors_cat1):
            self.gen_style_cat1_(f'style_cat1_{i+1}', c)
            
        style_cat2 = "style_cat2"
        if style_cat2 not in wb.named_styles:
            style_cat2 = NamedStyle(
                name=style_cat2,
                font=Font(color="FF808080",italic=True),
                fill=PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid"),
                border=self.thin_border,
                alignment=Alignment(horizontal='center', vertical='center'),
            )
            wb.add_named_style(style_cat2)

    def gen_style_cat1_(self, name, color):
        if name not in self.wb.named_styles:
            style = NamedStyle(
                name=name,
                font=Font(color="FF000000"),
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                border=self.thin_border,
                alignment=Alignment(horizontal='center', vertical='center'),
            )
            self.wb.add_named_style(style)
            