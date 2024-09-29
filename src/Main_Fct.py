from Mark_Sheet import SheetMarker
from Ext_PDF_extraction import PDFExtraction
from Ext_Write import WriteExtraction
from Real_Compute import MonthRealisation
from Real_Write import WriteRealisation
from Validation import Valid
from Styles import StyleCell
from datetime import datetime

def extraction(ws,mks,path):
    # Extract
    extraction = PDFExtraction(path)
    # Write
    w = WriteExtraction(ws,extraction,mks)
    # Valid
    if w.is_valid:
        ws.cell(row=mks.B_VALID_EXT_PATH[0],column=mks.B_VALID_EXT_PATH[1]).value = PATH_EXTRACTION
        ws.cell(row=mks.B_VALID_EXT_DATE[0],column=mks.B_VALID_EXT_DATE[1]).value = datetime.now()
        
def realisation(ws,mks):
    # Realise
    realisation = MonthRealisation(ws,mks)
    # Write
    WriteRealisation(ws,mks,realisation)
    # Valid
    ws.cell(row=mks.B_VALID_REAL_DATE[0],column=mks.B_VALID_REAL_DATE[1]).value = datetime.now()

if __name__ == "__main__":
    import openpyxl
    path_excel = r"D:\One Drive\OneDrive\Bureau\Classeur1.xlsx"
    workbook = openpyxl.load_workbook(path_excel)
    StyleCell(workbook)
    ws = workbook.active
    mks = SheetMarker(ws)
    path_extraction = r"D:\One Drive\OneDrive\Bureau\relevejuin.pdf"
    extraction(ws,mks,path_extraction)
    # realisation(ws,mks)
    Valid(ws,mks)
    
    workbook.save(path_excel)