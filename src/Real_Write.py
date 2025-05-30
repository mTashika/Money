import Const as C
import Const_txt_excel as CTE
from Const import REAL_LINE_SPACE,REAL_LINE_SPACE_SIGN,REAL_SPACE_INC_LOSS_2_DETAIL
from openpyxl.styles import Side, Border,Alignment
from Tools import clear_zone,generate_table,unmerge_cells_by_coords,check_cell_value
from Protection import unprotect_range
from openpyxl.utils import get_column_letter

class IncomeLoss():
    def __init__(self,name,ex,re):
        self.name = name
        self.ex = ex
        self.re = re
        
class WriteRealisation():
    def __init__(self,ws,mks,realisation):
        self.ws = ws
        self.mks = mks
        self.realisation = realisation
        self.money_type = '€'
        self.in_list = []
        self.los_list = []
        self.get_income_loss()
        self.clear_real()
        self.print_real_income()
        self.print_real_loss()
        self.detail_merge()
        self.print_real_bilan()
        self.print_tot()
        self.print_cat()
        self.set_protection()

    def get_income_loss(self):
        # Income
        for row in range(self.mks.B_IN_ST_LINE,self.mks.B_DETAIL_LINE_ST-1):
            name = self.ws.cell(row=row,column=self.mks.B_IN_NAME_COL).value
            ex = self.ws.cell(row=row,column=self.mks.B_IN_EX_COL).value
            re = self.ws.cell(row=row,column=self.mks.B_IN_RE_COL).value
            if check_cell_value(name)==0 and (check_cell_value(ex) in [1,2] or check_cell_value(re) in [1,2]):
                inc = IncomeLoss(name,ex,None)
                self.in_list.append(inc)
        # Loss
        for row in range(self.mks.B_LOS_ST_LINE,self.mks.B_DETAIL_LINE_ST-1):
            name = self.ws.cell(row=row,column=self.mks.B_LOS_NAME_COL).value
            ex = self.ws.cell(row=row,column=self.mks.B_LOS_EX_COL).value
            re = self.ws.cell(row=row,column=self.mks.B_LOS_RE_COL).value
            if check_cell_value(name)==0 and (check_cell_value(ex) in [1,2] or check_cell_value(re) in [1,2]):
                los = IncomeLoss(name,ex,None)
                self.los_list.append(los)
        
    def clear_real(self):
        clear_zone(self.ws,self.mks.B_IN_ST_LINE,self.mks.B_DETAIL_LINE_ST-2,self.mks.B_IN_NAME_COL,self.mks.B_IN_RE_COL)
        clear_zone(self.ws,self.mks.B_LOS_ST_LINE,self.mks.B_DETAIL_LINE_ST-2,self.mks.B_LOS_NAME_COL,self.mks.B_LOS_RE_COL)
        clear_zone(self.ws,self.mks.B_DETAIL_LINE_ST,self.ws.max_row+1,self.mks.B_REAL_COL_ST,self.mks.B_REAL_COL_ED)
        unmerge_cells_by_coords(self.ws,self.mks.B_DETAIL_LINE_ST-1,self.mks.B_ID_C1_COL,self.mks.B_DETAIL_LINE_ST-1,self.mks.B_REAL_COL_ED)
        while ((self.mks.B_DETAIL_LINE_ST > (self.mks.B_IN_ST_LINE + REAL_SPACE_INC_LOSS_2_DETAIL + 2)) and 
                (self.mks.B_DETAIL_LINE_ST > (self.mks.B_LOS_ST_LINE + REAL_SPACE_INC_LOSS_2_DETAIL + 2))):
            self.ws.delete_rows(self.mks.B_IN_ST_LINE+1)
            self.mks.B_DETAIL_LINE_ST-=1
        self.mks.B_in_ed_line = self.mks.B_IN_ST_LINE
        self.mks.B_los_ed_line = self.mks.B_LOS_ST_LINE
            
    def print_tot(self):
        # total name
        r1 = self.mks.B_DETAIL_LINE_ST
        c1 = self.mks.B_REAL_COL_ST
        self.ws.cell(r1,c1).value   = CTE.REAL_TOT
        self.ws.cell(r1,c1+1).value = CTE.REAL_TOT_EXP
        self.ws.cell(r1,c1+2).value = CTE.REAL_TOT_INC
        self.ws.cell(r1,c1+3).value = CTE.REAL_TOT_REAL
        self.ws.cell(r1,c1+4).value = CTE.REAL_TOT_REAL_EXP
        self.ws.cell(r1,c1+5).value = CTE.REAL_TOT_REAL_INC
        s = "style_title"
        self.ws.cell(r1,c1).style   = s
        self.ws.cell(r1,c1+1).style = s
        self.ws.cell(r1,c1+2).style = s
        self.ws.cell(r1,c1+3).style = s
        self.ws.cell(r1,c1+4).style = s
        self.ws.cell(r1,c1+5).style = s
        # total value
        self.ws.cell(r1+1,c1).value   = self.realisation.tot
        self.ws.cell(r1+1,c1+1).value = self.realisation.tot_ex
        self.ws.cell(r1+1,c1+2).value = self.realisation.tot_inc
        self.ws.cell(r1+1,c1+3).value = self.realisation.tot_real
        self.ws.cell(r1+1,c1+4).value = self.realisation.tot_ex_real
        self.ws.cell(r1+1,c1+5).value = self.realisation.tot_inc_real
        s = "style_title_val"
        self.ws.cell(r1+1,c1).style   = s
        self.ws.cell(r1+1,c1+1).style = s
        self.ws.cell(r1+1,c1+2).style = s
        self.ws.cell(r1+1,c1+3).style = s
        self.ws.cell(r1+1,c1+4).style = s
        self.ws.cell(r1+1,c1+5).style = s
        # total border
        self.ws.cell(r1,c1+3).border   = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thin'))
        self.ws.cell(r1,c1+4).border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thin'))
        self.ws.cell(r1,c1+5).border   = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thin'))
        self.ws.cell(r1+1,c1+3).border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thick'))
        self.ws.cell(r1+1,c1+4).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thick'))
        self.ws.cell(r1+1,c1+5).border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thin'),bottom=Side(style='thick'))
    
    def print_cat(self):
        c1 = self.mks.B_REAL_COL_ST
        row_c = self.mks.B_DETAIL_LINE_ST + 3
        round_p = 2
        style_cat2 = "style_cat2"
        tab_color_cat1 = generate_table(len(self.realisation.cats1),8)
        
        for i,cat1 in enumerate(self.realisation.cats1):
            if i!=0:
                if cat1.tot*self.ws.cell(row_c-1,c1+2).value < 0: # signe change
                    row_c += REAL_LINE_SPACE_SIGN
            row_st = row_c
            style_cat1_tot = f"style_cat1_{tab_color_cat1[i]}"
            # write cat1 name
            self.ws.cell(row_c,c1).value = cat1.name
            self.ws.cell(row_c,c1).style = style_cat1_tot
            
            # write all cat2
            if cat1.nb_cat2 == 0:
                self.ws.cell(row_c,c1+1).value = CTE.REAL_C1_TOT
            elif cat1.nb_cat2 == 1:
                self.ws.cell(row_c,c1+1).value = cat1.cats2[0].name
            else:
                # write all cat2 and increase row_c
                for cat2 in cat1.cats2:
                    if cat2.name is not None:
                        self.ws.cell(row_c,c1+1).value = cat2.name
                        self.ws.cell(row_c,c1+2).value = cat2.tot
                        self.ws.cell(row_c,c1+1).style = style_cat2
                        self.ws.cell(row_c,c1+2).style = style_cat2
                        
                        self.ws.cell(row_c,c1+1).alignment = Alignment(horizontal='left', vertical='center')
                        
                        self.ws.cell(row_c,c1+2).number_format = f'0"{self.money_type}"'
                        self.ws.cell(row_c,c1+2).alignment = Alignment(horizontal='center', vertical='center')
                        if cat1.disp_p:
                            cell = self.ws.cell(row_c,c1+3)
                            cell.value = round(cat2.tot_p,round_p)
                            cell.style = style_cat2
                            cell.number_format = '0%'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            
                        row_c+=1
                # merge cat1 title
                self.ws.merge_cells(start_row=row_st, start_column=c1,
                            end_row=row_c, end_column=c1)
                # write total line
                self.ws.cell(row_c,c1+1).value = CTE.REAL_C1_TOT
                
            # write cat1 tot and tot_p
            self.ws.cell(row_c,c1+2).value = cat1.tot
            if cat1.disp_p:
                cell = self.ws.cell(row_c,c1+3)
                cell.value = round(cat1.tot_p,round_p)
                cell.style = style_cat1_tot
                cell.number_format = '0%'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            self.ws.cell(row_c,c1+1).style = style_cat1_tot
            self.ws.cell(row_c,c1+2).style = style_cat1_tot
            self.ws.cell(row_c,c1+2).number_format = f'0"{self.money_type}"'
            self.ws.cell(row_c,c1+2).alignment = Alignment(horizontal='center', vertical='center')
            
            row_c += 1 + REAL_LINE_SPACE

    def print_real_income(self):
        for cat1 in self.realisation.cats1:
            if cat1.tot>0:
                f_ok = False
                for income in self.in_list:
                    if income.name == cat1.name:
                        income.re = cat1.tot
                        f_ok = True
                        break
                if not f_ok:
                    self.in_list.append(IncomeLoss(cat1.name,None,cat1.tot))
                    
        for income in self.in_list:
            if check_cell_value(income.name)==0 and(check_cell_value(income.ex) in [1,2] or check_cell_value(income.re) in [1,2]):
                while (self.mks.B_DETAIL_LINE_ST <= (self.mks.B_in_ed_line + REAL_SPACE_INC_LOSS_2_DETAIL + 1)):
                    self.ws.insert_rows(self.mks.B_in_ed_line+1)
                    self.mks.B_DETAIL_LINE_ST+=1
                self.ws.cell(self.mks.B_in_ed_line,self.mks.B_IN_NAME_COL).value = income.name
                self.ws.cell(self.mks.B_in_ed_line,self.mks.B_IN_RE_COL).value = income.re
                self.ws.cell(self.mks.B_in_ed_line,self.mks.B_IN_EX_COL).value = income.ex
                self.mks.B_in_ed_line+=1
        self.maj_formule_income()
    
    def maj_formule_income(self):
        self.ws.cell(self.mks.B_IN_TOT_EX[0],self.mks.B_IN_TOT_EX[1]).value = f'=SUM({get_column_letter(self.mks.B_IN_EX_COL)}{self.mks.B_IN_ST_LINE}:{get_column_letter(self.mks.B_IN_EX_COL)}{self.mks.B_in_ed_line})'
        self.ws.cell(self.mks.B_IN_TOT_REAL[0],self.mks.B_IN_TOT_REAL[1]).value = f'=SUM({get_column_letter(self.mks.B_IN_RE_COL)}{self.mks.B_IN_ST_LINE}:{get_column_letter(self.mks.B_IN_RE_COL)}{self.mks.B_in_ed_line})'
        
    def print_real_loss(self):
        for cat1 in self.realisation.cats1:
            if cat1.tot<0:
                f_ok = False
                for loss in self.los_list:
                    if loss.name == cat1.name:
                        loss.re = cat1.tot
                        f_ok = True
                        break
                if not f_ok:
                    self.los_list.append(IncomeLoss(cat1.name,None,cat1.tot))
        
        for loss in self.los_list:
            if check_cell_value(loss.name)==0 and(check_cell_value(loss.ex) in [1,2] or check_cell_value(loss.re) in [1,2]):
                while (self.mks.B_DETAIL_LINE_ST <= (self.mks.B_los_ed_line + REAL_SPACE_INC_LOSS_2_DETAIL + 1)):
                    self.ws.insert_rows(self.mks.B_los_ed_line+1)
                    self.mks.B_DETAIL_LINE_ST+=1
                self.ws.cell(self.mks.B_los_ed_line,self.mks.B_LOS_NAME_COL).value = loss.name
                self.ws.cell(self.mks.B_los_ed_line,self.mks.B_LOS_RE_COL).value = loss.re
                self.ws.cell(self.mks.B_los_ed_line,self.mks.B_LOS_EX_COL).value = loss.ex
                self.mks.B_los_ed_line+=1
        self.maj_formule_loss()

    def maj_formule_loss(self):
        self.ws.cell(self.mks.B_LOS_TOT_EX[0],self.mks.B_LOS_TOT_EX[1]).value = f'=SUM({get_column_letter(self.mks.B_LOS_EX_COL)}{self.mks.B_LOS_ST_LINE}:{get_column_letter(self.mks.B_LOS_EX_COL)}{self.mks.B_los_ed_line})'
        self.ws.cell(self.mks.B_LOS_TOT_REAL[0],self.mks.B_LOS_TOT_REAL[1]).value = f'=SUM({get_column_letter(self.mks.B_LOS_RE_COL)}{self.mks.B_LOS_ST_LINE}:{get_column_letter(self.mks.B_LOS_RE_COL)}{self.mks.B_los_ed_line})'
    
    def detail_merge(self):
        self.ws.merge_cells(start_row=self.mks.B_DETAIL_LINE_ST-1, start_column=self.mks.B_ID_C1_COL,
                            end_row=self.mks.B_DETAIL_LINE_ST-1, end_column=self.mks.B_REAL_COL_ED)
        
    def extend_income_loss(self,type):
        #unmerge detail
        unmerge_cells_by_coords(self.ws,self.mks.B_DETAIL_LINE_ST-1,self.mks.B_ID_C1_COL,self.mks.B_DETAIL_LINE_ST-1,self.mks.B_REAL_COL_ED)
        if type == 'in':
            self.ws.insert_rows(self.mks.B_in_ed_line+1)
        elif type == 'loss':
            self.ws.insert_rows(self.mks.B_los_ed_line+1)
        self.mks.B_DETAIL_LINE_ST+=1
        # merge detail
        self.ws.merge_cells(start_row=self.mks.B_DETAIL_LINE_ST-1, start_column=self.mks.B_ID_C1_COL,
                            end_row=self.mks.B_DETAIL_LINE_ST-1, end_column=self.mks.B_REAL_COL_ED)

    def print_real_bilan(self):
        self.ws.cell(self.mks.B_BILAN_ED_SOLDRE[0],self.mks.B_BILAN_ED_SOLDRE[1]).value = f'={get_column_letter(self.mks.B_IN_TOT_REAL[1])}{self.mks.B_IN_TOT_REAL[0]}+{get_column_letter(self.mks.B_LOS_TOT_REAL[1])}{self.mks.B_LOS_TOT_REAL[0]}+{get_column_letter(self.mks.B_BILAN_ST_SOLD[1])}{self.mks.B_BILAN_ST_SOLD[0]}'

    def set_protection(self):
        unprotect_range(self.ws,self.mks.B_IN_ST_LINE,self.mks.B_DETAIL_LINE_ST-2,self.mks.B_IN_NAME_COL,self.mks.B_REAL_COL_ED)
        unprotect_range(self.ws,self.mks.B_LOS_ST_LINE,self.mks.B_DETAIL_LINE_ST-2,self.mks.B_LOS_NAME_COL,self.mks.B_REAL_COL_ED)