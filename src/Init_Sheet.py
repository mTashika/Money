import Const_txt_excel as CET
import Const_Balise_Excel as CEB
from openpyxl.styles import Font,Alignment,PatternFill,Side,Border
from Protection import set_protection

FORMAT_COMPTA = "0.00 €"


class InitSheet:
    def __init__(self, ws):
        self.ws = ws
        self.set_big_border()
        self.set_line_height()
        self.write_start()
        self.write_bilan()
        self.write_income()
        self.write_loss()
        self.write_detail()
        self.write_tab()
        self.write_validity()
        self.set_column_width()
        set_protection(self.ws)
        # self.set_protection()

    def set_big_border(self):
        """ Dotted Border """
        for col in range(1, 16):
            cell = self.ws.cell(row=10, column=col)
            cell.border =  Border(bottom=Side(border_style='dotted', color='000000'))
        for row in range(1, 10):
            cell = self.ws.cell(row=row, column=16)
            cell.border =  Border(right=Side(border_style='dotted', color='000000'))
        cell = self.ws.cell(row=10, column=16)
        cell.border =  Border(bottom=Side(border_style='dotted', color='000000'),right=Side(border_style='dotted', color='000000'))

    def set_column_width(self):
        self.ws.column_dimensions['A'].width = 15
        self.ws.column_dimensions['B'].width = 17
        self.ws.column_dimensions['C'].width = 13
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 55
        self.ws.column_dimensions['F'].width = 12
        self.ws.column_dimensions['G'].width = 3
        self.ws.column_dimensions['H'].width = 19
        self.ws.column_dimensions['I'].width = 19
        self.ws.column_dimensions['J'].width = 19
        self.ws.column_dimensions['K'].width = 19
        self.ws.column_dimensions['L'].width = 5
        self.ws.column_dimensions['M'].width = 19
        self.ws.column_dimensions['N'].width = 19
        self.ws.column_dimensions['O'].width = 19
        self.ws.column_dimensions['P'].width = 19
        
    def set_line_height(self):
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 20
        self.ws.row_dimensions[3].height = 30
        self.ws.row_dimensions[4].height = 24
        self.ws.row_dimensions[5].height = 21
        self.ws.row_dimensions[6].height = 15
        self.ws.row_dimensions[7].height = 21
        self.ws.row_dimensions[8].height = 20
        self.ws.row_dimensions[9].height = 15
        self.ws.row_dimensions[10].height = 20
        self.ws.row_dimensions[11].height = 20
        self.ws.row_dimensions[12].height = 20
        self.ws.row_dimensions[13].height = 20

    def write_start(self):
        cell_date_b = self.ws.cell(1, 1)
        cell_start_b = self.ws.cell(2, 1)
        cell_sold_st = self.ws.cell(2,2)
        cell_date_txt = self.ws.cell(1,2) 
        
        font_style = Font(size=14, bold=True)
        cell_date_b.font = font_style
        cell_start_b.font = font_style
        cell_sold_st.font = font_style
        cell_date_txt.font = Font(size=13, bold=True)
        
        cell_sold_st.number_format = FORMAT_COMPTA
        
        cell_date_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_start_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_sold_st.alignment = Alignment(horizontal='right', vertical='center')
        cell_date_txt.alignment = Alignment(horizontal='center', vertical='center')

        cell_date_b.value = CEB.BILAN_DATE
        cell_date_b.number_format = f';;;"{CET.INIT_DATE}"'
        cell_start_b.value = CEB.BILAN_ST_SOLD
        cell_start_b.number_format = f';;;"{CET.INIT_SOLD_ST}"'
        
    def write_bilan(self):  
        cell_title_b  = self.ws.cell(4,1)
        cell_title_ex = self.ws.cell(4,2)
        cell_title_re = self.ws.cell(4,3)
        cell_val_ex   = self.ws.cell(5,2)
        cell_val_re   = self.ws.cell(5,3)
        cell_valid_b  = self.ws.cell(3,4)
        cell_valid_note  = self.ws.cell(2,4)
        
        cell_title_b.font = Font(size=16, bold=True)
        cell_title_ex.font = Font(size=16, bold=True)
        cell_title_re.font = Font(size=16, bold=True)
        cell_val_ex.font = Font(size=14)
        cell_val_re.font = Font(size=14)
        cell_valid_b.font = Font(size=14)
        cell_valid_note.font = Font(size=9, color="FFFFFF")
        
        cell_title_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_ex.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_re.alignment = Alignment(horizontal='center', vertical='center')
        cell_val_ex.alignment = Alignment(horizontal='right', vertical='center')
        cell_val_re.alignment = Alignment(horizontal='right', vertical='center')
        cell_valid_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_valid_note.alignment = Alignment(horizontal='left', vertical='center')
        
        cell_val_ex.number_format = FORMAT_COMPTA
        cell_val_re.number_format = FORMAT_COMPTA
        
        fill_color1 = PatternFill(start_color='83CCEB', end_color='83CCEB', fill_type='solid')
        fill_color2 = PatternFill(start_color='C0E6F5', end_color='C0E6F5', fill_type='solid')
        cell_title_b.fill = fill_color1
        cell_title_ex.fill = fill_color1
        cell_title_re.fill = fill_color1
        cell_val_ex.fill = fill_color2
        cell_val_re.fill = fill_color2

        cell_title_b.border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_ex.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thin'))
        cell_title_re.border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thin'))
        cell_val_ex.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thick'))
        cell_val_re.border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thin'),bottom=Side(style='thick'))

        self.ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
        
        cell_title_b.value = CEB.BILAN_ED_SOLD
        cell_title_b.number_format = f';;;"{CET.INIT_TITLE_BILAN}"'
        cell_valid_b.value = ""
        cell_valid_b.number_format = f';;;"{CET.INIT_VALID}"'
        cell_valid_note.value = CET.INIT_VALID_TOT_NOTE
        
        
        cell_title_ex.value = CET.INIT_TITLE_EXPECTED
        cell_title_re.value = CET.INIT_TITLE_REAL
        cell_val_ex.value = '=B2+B8+F8'

    def write_income(self):
        cell_title_b = self.ws.cell(7,1)
        cell_title_tot = self.ws.cell(8,1)
        cell_val_ex = self.ws.cell(8,2)
        cell_val_re = self.ws.cell(8,3)
        cell_title_name = self.ws.cell(10,1)
        cell_title_ex = self.ws.cell(10,2)
        cell_title_re = self.ws.cell(10,3)

        cell_title_b.font = Font(size=16, bold=True)
        cell_title_tot.font = Font(size=16, bold=True)
        cell_title_name.font = Font(size=14)
        cell_title_ex.font = Font(size=14)
        cell_title_re.font = Font(size=14)
        cell_val_ex.font = Font(size=14)
        cell_val_re.font = Font(size=14)
        
        cell_title_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_tot.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_name.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_ex.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_re.alignment = Alignment(horizontal='center', vertical='center')
        cell_val_ex.alignment = Alignment(horizontal='right', vertical='center')
        cell_val_re.alignment = Alignment(horizontal='right', vertical='center')
        
        cell_val_ex.number_format = FORMAT_COMPTA
        cell_val_re.number_format = FORMAT_COMPTA
        
        fill_color = PatternFill(start_color='C1F0C8', end_color='C1F0C8', fill_type='solid')
        cell_title_b.fill = PatternFill(start_color='83E28E', end_color='83E28E', fill_type='solid')
        cell_title_tot.fill = fill_color
        cell_title_name.fill = fill_color
        cell_title_ex.fill = fill_color
        cell_title_re.fill = fill_color
        cell_val_ex.fill = fill_color
        cell_val_re.fill = fill_color

        cell_title_b.border = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_tot.border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_name.border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_ex.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_re.border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_val_ex.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_val_re.border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))

        self.ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=3)
        
        cell_val_ex.value = "=SUM(B11:B25)"
        cell_title_b.value = CEB.INCOME
        cell_title_b.number_format = f';;;"{CET.INIT_TITLE_IN}"'
        cell_title_tot.value = CET.INIT_TITLE_TOT
        cell_title_name.value = CET.INIT_TITLE_NAME
        cell_title_ex.value = CET.INIT_TITLE_EXPECTED
        cell_title_re.value = CET.INIT_TITLE_REAL
    
    def write_loss(self):
        cell_title_b = self.ws.cell(7,5)
        cell_title_tot = self.ws.cell(8,5)
        cell_val_ex = self.ws.cell(8,6)
        cell_val_re = self.ws.cell(8,8)
        cell_title_name = self.ws.cell(10,5)
        cell_title_ex = self.ws.cell(10,6)
        cell_title_re = self.ws.cell(10,8)

        cell_title_b.font = Font(size=16, bold=True)
        cell_title_tot.font = Font(size=16, bold=True)
        cell_title_name.font = Font(size=14)
        cell_title_ex.font = Font(size=14)
        cell_title_re.font = Font(size=14)
        cell_val_ex.font = Font(size=14)
        cell_val_re.font = Font(size=14)
        
        cell_title_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_tot.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_name.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_ex.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_re.alignment = Alignment(horizontal='center', vertical='center')
        cell_val_ex.alignment = Alignment(horizontal='right', vertical='center')
        cell_val_re.alignment = Alignment(horizontal='right', vertical='center')
        
        cell_val_ex.number_format = FORMAT_COMPTA
        cell_val_re.number_format = FORMAT_COMPTA
        
        fill_color = PatternFill(start_color='FBE2D5', end_color='FBE2D5', fill_type='solid')
        cell_title_b.fill = PatternFill(start_color='F7C7AC', end_color='F7C7AC', fill_type='solid')
        cell_title_tot.fill = fill_color
        cell_title_name.fill = fill_color
        cell_title_ex.fill = fill_color
        cell_title_re.fill = fill_color
        cell_val_ex.fill = fill_color
        cell_val_re.fill = fill_color

        cell_title_b.border = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_tot.border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_name.border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_ex.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_title_re.border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_val_ex.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_val_re.border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))

        self.ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=8)
        self.ws.merge_cells(start_row=8, start_column=6, end_row=8, end_column=7)
        self.ws.merge_cells(start_row=10, start_column=6, end_row=10, end_column=7)
        
        cell_val_ex.value = "=SUM(F11:F25)"
        cell_title_b.value = CEB.LOSS
        cell_title_b.number_format = f';;;"{CET.INIT_TITLE_LOS}"'
        cell_title_tot.value = CET.INIT_TITLE_TOT
        cell_title_name.value = CET.INIT_TITLE_NAME
        cell_title_ex.value = CET.INIT_TITLE_EXPECTED
        cell_title_re.value = CET.INIT_TITLE_REAL

        for i in range(11,26):
            self.ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)

    def write_detail(self):
        cell_detail_b = self.ws.cell(26,1)
        cell_est_sold = self.ws.cell(29,1)

        cell_detail_b.font = Font(size=11)
        cell_est_sold.font = Font(size=11)
        
        cell_detail_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_est_sold.alignment = Alignment(horizontal='center', vertical='center')

        for i in range(1,16):
            self.ws.cell(26,i).border = Border(top=Side(style='thick'),bottom=Side(style='thick'))
        self.ws.cell(26,16).border = Border(top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'))

        self.ws.merge_cells(start_row=26, start_column=1, end_row=26, end_column=16)
        cell_detail_b.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        cell_detail_b.value = CEB.DETAIL
        cell_detail_b.number_format = f';;;""'
        cell_est_sold.value = CEB.EXT_VAL
    
    def write_tab(self):
        cell_cat1   = self.ws.cell(27,1)
        cell_cat2   = self.ws.cell(27,2)
        cell_intern = self.ws.cell(27,3)
        cell_share  = self.ws.cell(27,4)
        cell_name   = self.ws.cell(27,5)
        cell_value  = self.ws.cell(27,6)


        fill_color = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        cell_cat1.fill = fill_color
        cell_cat2.fill = fill_color
        cell_intern.fill = fill_color
        cell_share.fill = fill_color
        cell_name.fill = fill_color
        cell_value.fill = fill_color

        cell_cat1.font = Font(size=14, color="FFFFFF")
        cell_cat2.font = Font(size=14, color="FFFFFF")
        cell_intern.font = Font(size=14, color="FFFFFF")
        cell_share.font = Font(size=14, color="FFFFFF")
        cell_name.font = Font(size=14, color="FFFFFF")
        cell_value.font = Font(size=14, color="FFFFFF")

        cell_cat1.alignment = Alignment(horizontal='center', vertical='center')
        cell_cat2.alignment = Alignment(horizontal='center', vertical='center')
        cell_intern.alignment = Alignment(horizontal='center', vertical='center')
        cell_share.alignment = Alignment(horizontal='center', vertical='center')
        cell_name.alignment = Alignment(horizontal='center', vertical='center')
        cell_value.alignment = Alignment(horizontal='center', vertical='center')

        cell_cat1.border = Border(left=Side(style='thick'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_cat2.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_intern.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_share.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_name.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
        cell_value.border = Border(left=Side(style='thin'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))

        cell_cat1.value = CET.INIT_TAB_COL_CAT1
        cell_cat2.value = CET.INIT_TAB_COL_CAT2
        cell_intern.value = CET.INIT_TAB_COL_TYPE
        cell_share.value = CET.INIT_TAB_COL_REFUND
        cell_name.value = CET.INIT_TAB_COL_NAME
        cell_value.value = CET.INIT_TAB_COL_VALUE

    def write_validity(self):
        cell_title_b       = self.ws.cell(1,8)
        cell_title_ex1     = self.ws.cell(2,8)
        cell_title_ex2     = self.ws.cell(3,8)
        cell_title_id      = self.ws.cell(4,8)
        cell_title_re      = self.ws.cell(5,8)
        cell_title_path    = self.ws.cell(2,9)
        cell_title_date_ex = self.ws.cell(3,9)
        cell_title_date_id = self.ws.cell(4,9)
        cell_title_date_re = self.ws.cell(5,9)
        cell_val_ex_path   = self.ws.cell(2,10)
        cell_val_ex_date   = self.ws.cell(3,10)
        cell_val_id_date   = self.ws.cell(4,10)
        cell_val_re_date   = self.ws.cell(5,10)
        
        cell_title_b.font = Font(size=9, color="FFFFFF")
        cell_title_ex1.font = Font(size=9, color="FFFFFF")
        cell_title_ex2.font = Font(size=9, color="FFFFFF")
        cell_title_id.font = Font(size=9, color="FFFFFF")
        cell_title_re.font = Font(size=9, color="FFFFFF")
        cell_title_path.font = Font(size=9, color="FFFFFF")
        cell_title_date_ex.font = Font(size=9, color="FFFFFF")
        cell_title_date_id.font = Font(size=9, color="FFFFFF")
        cell_title_date_re.font = Font(size=9, color="FFFFFF")
        cell_val_ex_path.font = Font(size=9, color="FFFFFF")
        cell_val_ex_date.font = Font(size=9, color="FFFFFF")
        cell_val_id_date.font = Font(size=9, color="FFFFFF")
        cell_val_re_date.font = Font(size=9, color="FFFFFF")
        
        cell_title_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_title_ex1.alignment = Alignment(horizontal='left', vertical='center')
        cell_title_ex2.alignment = Alignment(horizontal='left', vertical='center')
        cell_title_id.alignment = Alignment(horizontal='left', vertical='center')
        cell_title_re.alignment = Alignment(horizontal='left', vertical='center')
        cell_title_path.alignment = Alignment(horizontal='left', vertical='center')
        cell_title_date_ex.alignment = Alignment(horizontal='left', vertical='center')
        cell_title_date_id.alignment = Alignment(horizontal='left', vertical='center')
        cell_title_date_re.alignment = Alignment(horizontal='left', vertical='center')
        cell_val_ex_path.alignment = Alignment(horizontal='left', vertical='center')
        cell_val_ex_date.alignment = Alignment(horizontal='left', vertical='center')
        cell_val_id_date.alignment = Alignment(horizontal='left', vertical='center')
        cell_val_re_date.alignment = Alignment(horizontal='left', vertical='center')
        
        self.ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
        
        cell_title_b.value = CEB.VALIDITY
        cell_title_b.number_format = f';;;"{CET.INIT_TITLE_VALIDITY}"'
        cell_title_ex1.value = CET.INIT_TITLE_EXTRACTION
        cell_title_ex2.value = CET.INIT_TITLE_EXTRACTION
        cell_title_id.value = CET.INIT_TITLE_IDENTIFICATION
        cell_title_re.value = CET.INIT_TITLE_REALISATION
        cell_title_path.value = CET.INIT_TITLE_PATH
        cell_title_date_ex.value = CET.INIT_TITLE_DATE
        cell_title_date_id.value = CET.INIT_TITLE_DATE
        cell_title_date_re.value = CET.INIT_TITLE_DATE
        
if __name__ == "__main__":
    import openpyxl
    from Styles import StyleCell
    
    path = r"C:\Users\mcast\OneDrive\Bureau\Classeur1.xlsx"
    workbook = openpyxl.load_workbook(path)
    StyleCell(workbook)
    ws = workbook.active
    InitSheet(ws)
    workbook.save(path)
    