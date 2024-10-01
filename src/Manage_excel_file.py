import os
import openpyxl
from tkinter import messagebox
from Init_Workbook import InitWorkbook
from Init_Sheet import InitSheet
from Protection import protect_wb
TYPE_EXTRACTION = "Extraction"
TYPE_REALISATION = "Realisation"

class ManageExcelFile:
    def __init__(self,folder_path,year,month,managment_type):
        self.folder_path = folder_path
        self.file_name = f'{year}.xlsx'
        self.ws_name = month
        self.managment_type = managment_type
        
        self.file_path = os.path.join(self.folder_path, self.file_name)
        
        self.wb_valid = False
        self.ws_valid = False
        
        self.wb_creat = False
        self.ws_creat = False
        self.wb = None
        self.ws = None
        
        self.get_workbook_and_sheet()
        
    def get_workbook_and_sheet(self):
        wb_exists, sheet_exists = self.check_excel_file_and_sheet()
        if not wb_exists:
            if self.managment_type == TYPE_EXTRACTION:
                self.wb = openpyxl.Workbook()   # Create the Workbook
                protect_wb(self.wb)
                InitWorkbook(self.wb)           # Initialise the Workbook
                if self.wb:
                    self.wb_creat = True
                    self.wb_valid = True
        else:
            self.wb = openpyxl.load_workbook(self.file_path)
            self.wb_valid = True if self.wb else False
        if not sheet_exists:
            if self.managment_type == TYPE_EXTRACTION:
                self.ws = self.wb.create_sheet(title=self.ws_name)
                InitSheet(self.ws)
                if self.ws:
                    self.ws_creat = True
                    self.ws_valid = True
        else:
            self.ws = self.wb[self.ws_name]
            self.ws_valid = True if self.ws else False
        
    def check_excel_file_and_sheet(self):
        """
        Vérifie si un fichier Excel existe, est accessible, et si un onglet spécifique existe dans ce fichier.
        :param self.file_path: Chemin du fichier Excel.
        :param self.ws_name: Nom de l'onglet à vérifier.
        :return: Tuple (wb_exists, sheet_exists) :
                - wb_exists: True si le fichier existe et est accessible, False sinon.
                - sheet_exists: True si l'onglet existe et est accessible, False sinon.
        """
        try:
            wb_exists = os.path.isfile(self.file_path) and os.access(self.file_path, os.R_OK)
            sheet_exists = False
            if wb_exists:
                try:
                    workbook = openpyxl.load_workbook(self.file_path, read_only=True)
                    if self.ws_name in workbook.sheetnames:
                        sheet_exists = True
                except Exception as e:
                    print(f"Une erreur s'est produite lors de l'accès au fichier ou à l'onglet: {e}")
            return wb_exists, sheet_exists
        except PermissionError:
            raise PermissionError(f"Permission refusée: Impossible d'accéder au fichier '{self.file_path}'. Veuillez vérifier que le fichier n'est pas ouvert dans une autre application.")
            
        except FileNotFoundError:
            raise FileNotFoundError(f"Fichier non trouvé: Impossible de localiser le chemin spécifié '{self.file_path}'.")
            
        except Exception as e:
            raise Exception(f"Une erreur s'est produite: {e}")
        
    def get_wb_ws(self):
        # if not self.wb_valid:
        #     print('Workbook not valid')
        #     return None,None
        # if not self.ws_valid:
        #     print('Worksheet not valid')
        #     return None,None
        return self.wb,self.ws
    
    def save_wb(self):
        if self.wb_valid and self.ws_valid:
            try:
                self.wb.save(self.file_path)
                return 1
            except PermissionError:
                messagebox.showerror('Error','Permission denied\nTry to close the excel first')
                print('Workbook not saved')
                return -1
            except FileNotFoundError:
                messagebox.showerror("Error", "Invalid Excel Path !")
                print('Workbook not saved')
                return -1
        else:
            print('Workbook not saved')