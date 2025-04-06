import Const as C
import Const_Balise_Excel as BE
from openpyxl.utils import get_column_letter
from Tools import clear_zone
from Protection import set_protection

class WriteExtraction():
    def __init__(self,sheet,extraction,markers):
        self.extraction = extraction
        self.fo = self.extraction.financial_operations
        self.mks = markers
        self.ws = sheet
        self.supress_old()
        self.write()
        self.up_mrks()
        self.set_cells_format()
        self.write_validation()
        self.write_date()
        self.write_sold()
        self.write_income_loss()
        set_protection(self.ws,self.mks)

    def supress_old(self):
        clear_zone(self.ws,self.mks.B_DETAIL_LINE_ST,self.ws.max_row+1,self.mks.B_ID_C1_COL,self.mks.B_REAL_COL_ED)
        for row in range(self.mks.B_IN_ST_LINE,self.mks.B_in_ed_line+1):
            self.ws.cell(row=row,column=self.mks.B_IN_RE_COL).value = None
            if self.ws.cell(row=row,column=self.mks.B_IN_EX_COL).value is None:
                self.ws.cell(row=row,column=self.mks.B_IN_NAME_COL).value = None
        for row in range(self.mks.B_LOS_ST_LINE,self.mks.B_los_ed_line+1):
            self.ws.cell(row=row,column=self.mks.B_LOS_RE_COL).value = None
            if self.ws.cell(row=row,column=self.mks.B_LOS_EX_COL).value is None:
                self.ws.cell(row=row,column=self.mks.B_LOS_NAME_COL).value = None

    def write(self):
        for row,fo in enumerate(self.fo,start=self.mks.B_DETAIL_LINE_ST):
            self.ws.cell(row=row,column=self.mks.B_EXT_NAME_COL).value = fo.name
            self.ws.cell(row=row,column=self.mks.B_EXT_VALUE_COL).value = fo.value
            self.ws.cell(row=row,column=self.mks.B_ID_TYPE_COL).value = C.VALID_TYPE[0]
        self.mks.B_ext_ed_line = row + C.EXT_LINE_ADD
    
    def up_mrks(self):
        self.mks.b_ext_sold_ed_est = [self.mks.B_ext_ed_line+1,self.mks.B_ID_C2_COL]
        self.mks.b_ext_sold_ed_true = [self.mks.B_ext_ed_line+2,self.mks.B_ID_C2_COL]
        self.mks.b_ext_verif = [self.mks.b_ext_sold_ed_est[0],self.mks.B_EXT_NAME_COL]
    
    def set_cells_format(self):
        for row in range(self.mks.B_DETAIL_LINE_ST,self.mks.B_ext_ed_line+1):
            self.ws.cell(row=row,column=self.mks.B_EXT_NAME_COL).style = "Normal"
            self.ws.cell(row=row,column=self.mks.B_EXT_VALUE_COL).style = "Note"
            cell_cat1 = self.ws.cell(row=row,column=self.mks.B_ID_C1_COL)
            cell_cat2 =  self.ws.cell(row=row,column=self.mks.B_ID_C2_COL)
            cell_intern = self.ws.cell(row=row,column=self.mks.B_ID_TYPE_COL)
            cell_share =  self.ws.cell(row=row,column=self.mks.B_ID_REFUND_COL)
            cell_cat1.style = "style_cat1"
            cell_cat2.style = "style_cat2"
            cell_intern.style = "style_col1"
            cell_share.style = "style_col1"
    
    def write_validation(self):
        # estimated
        self.ws.cell(row=self.mks.b_ext_sold_ed_est[0],column=self.mks.b_ext_sold_ed_est[1]).value = self.extraction.sold_est
        cell_txt_est = self.ws.cell(row=self.mks.b_ext_sold_ed_est[0],column=self.mks.b_ext_sold_ed_est[1]-1)
        cell_txt_est.value = BE.EXT_VAL
        cell_txt_est.number_format  = f';;;"{C.EXT_VALID_TXT_SOLD_EST}"'
        # validation
        cell_val = self.ws.cell(row=self.mks.b_ext_verif[0],column=self.mks.b_ext_verif[1])
        cell_val.value = self.extraction.validation
        if cell_val.value:
            cell_val.style = "Good"
        else:
            cell_val.style = "Bad"
    
    def write_date(self):
        date = f"{self.extraction.date_st} -> {self.extraction.date_ed}"       
        self.ws.cell(row=self.mks.B_BILAN_DATE[0],column=self.mks.B_BILAN_DATE[1]).value = date

    def write_sold(self):
        self.ws.cell(row=self.mks.B_BILAN_ST_SOLD[0],column=self.mks.B_BILAN_ST_SOLD[1]).value = self.extraction.sold_st
        self.ws.cell(row=self.mks.B_BILAN_ED_SOLDRE[0],column=self.mks.B_BILAN_ED_SOLDRE[1]).value = self.extraction.sold_ed
    
    def write_income_loss(self):
        self.ws.cell(row=self.mks.B_IN_TOT_EX[0],column=self.mks.B_IN_TOT_EX[1]).value = f"=SUM({get_column_letter(self.mks.B_IN_EX_COL)}{self.mks.B_IN_ST_LINE}:{get_column_letter(self.mks.B_IN_EX_COL)}{self.mks.B_DETAIL_LINE_ST-2})"
        self.ws.cell(row=self.mks.B_LOS_TOT_EX[0],column=self.mks.B_LOS_TOT_EX[1]).value = f"=SUM({get_column_letter(self.mks.B_LOS_EX_COL)}{self.mks.B_LOS_ST_LINE}:{get_column_letter(self.mks.B_LOS_EX_COL)}{self.mks.B_DETAIL_LINE_ST-2})"
    
    def is_valid(self):
        return True