import customtkinter as ctk
import tkinter as tk
import os
from tkinter import filedialog,messagebox,PhotoImage,Entry
from Tools import normalize_string,get_current_month,get_current_year,set_all_data_validation,run_excel_path
from openpyxl import load_workbook
from Ext_PDF_extraction import PDFExtraction
from Manage_excel_file import ManageExcelFile
from Mark_Sheet import SheetMarker
from Ext_Write import WriteExtraction
from Real_Compute import MonthRealisation
from Real_Write import WriteRealisation
from Styles import StyleCell
from Const import TOOL_SHEET_NAME,MONTH_FR
from Validation import Valid
import threading
from PIL import Image, ImageTk
import sys
from Init_App import App
import time

GREY = "#999999"
WHITE = "white"

def ressource_path(rel_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path,rel_path)
        
class Main():
    def __init__(self):
        self.App = App()
        self.asset_path = ressource_path('asset')        
        self.wb, self.ws = None, None
        self.excel_file_path = None
        self.year, self.month = None, None
        self.managment = None
        self.is_processing = False
        self.running_thread = False
        
        # Variables pour stocker les chemins
        self.pdf_path = ctk.StringVar(value="")
        self.excel_dir_path = ctk.StringVar(value=r"C:\Users\mcast\OneDrive\Documents\.Perso\[4] Finance\Compte\Fiche")
        self.year_var = ctk.StringVar(value=get_current_year())
        self.year_var.trace_add("write", self.update_combo_box)
        self.month_var = ctk.StringVar(value=get_current_month())
        self.checkbox_extract_var = ctk.IntVar()
        self.checkbox_update_var = ctk.IntVar()
        self.checkbox_create_var = ctk.IntVar()
        self.year_create = ctk.StringVar(value=get_current_year())
        self.month_create = ctk.StringVar(value=get_current_month())


        self.App.entry_excel_dir.configure(textvariable=self.excel_dir_path)
        self.App.button_browse_excel_dir.configure(command=self.browse_excel_dir)
        self.App.entry_pdf.configure(textvariable=self.pdf_path)
        self.App.button_browse_pdf.configure(command=self.browse_pdf)
        self.App.checkbox_extract.configure(variable=self.checkbox_extract_var,command=lambda: self.on_checkbox_toggle(self.checkbox_extract_var))
        self.App.entry_year_create.configure(textvariable=self.year_create)
        self.App.entry_month_create.configure(variable=self.month_create,values=MONTH_FR)
        self.App.checkbox_create.configure(variable=self.checkbox_create_var,command=lambda: self.on_checkbox_toggle(self.checkbox_create_var))
        self.App.entry_year.configure(textvariable=self.year_var)
        self.App.entry_month.configure(variable=self.month_var)
        self.App.checkbox_update.configure(variable=self.checkbox_update_var,command=lambda: self.on_checkbox_toggle(self.checkbox_update_var))
        self.on_checkbox_toggle(self.checkbox_extract_var)

        self.App.button_ok.configure(command=self.ok)
        self.App.button_cancel.configure(command=self.cancel)
        self.App.close_button.configure(command=self.cancel)

        self.App.settings_button.configure(command=self.open_settings)
        self.App.help_button.configure(command=self.open_help)

        self.init_gif()

    def disable_update(self):
       self.App.label_update_title.configure(text_color=GREY)
       self.App.label_year.configure(text_color=GREY)
       self.App.entry_year.configure(state="disabled",text_color=GREY)
       self.App.label_month.configure(text_color=GREY)
       self.App.entry_month.configure(state="disabled",text_color=GREY)
       self.App.checkbox_update.configure(state="normal")
       self.checkbox_update_var.set(0)
    def enable_update(self):
        self.App.label_update_title.configure(text_color=WHITE)
        self.App.label_year.configure(state="normal",text_color=WHITE)
        self.App.entry_year.configure(state="normal",text_color=WHITE)
        self.App.label_month.configure(state="normal",text_color=WHITE)
        self.App.entry_month.configure(state="normal",text_color=WHITE)
        self.App.checkbox_update.configure(state="normal")
        self.checkbox_update_var.set(self.App.CHECK_UPDATE)
    def disable_create(self):
       self.App.label_create_title.configure(text_color=GREY)
       self.App.label_year_create.configure(text_color=GREY)
       self.App.entry_year_create.configure(state="disabled",text_color=GREY)
       self.App.label_month_create.configure(text_color=GREY)
       self.App.entry_month_create.configure(state="disabled",text_color=GREY)
       self.App.checkbox_create.configure(state="normal")
       self.checkbox_create_var.set(0)
    def enable_create(self):
        self.App.label_create_title.configure(text_color=WHITE)
        self.App.label_year_create.configure(state="normal",text_color=WHITE)
        self.App.entry_year_create.configure(state="normal",text_color=WHITE)
        self.App.label_month_create.configure(state="normal",text_color=WHITE)
        self.App.entry_month_create.configure(state="normal",text_color=WHITE)
        self.App.checkbox_create.configure(state="normal")
        self.checkbox_create_var.set(self.App.CHECK_CREATE)
    
    def disable_extract(self):
        self.App.label_extract_title.configure(text_color=GREY)
        self.App.entry_pdf.configure(state="disabled",text_color=GREY)
        self.App.button_browse_pdf.configure(state="disabled")
        self.App.checkbox_extract.configure(state="normal")
        self.checkbox_extract_var.set(0)
    def enable_extract(self):
        self.App.label_extract_title.configure(text_color=WHITE)
        self.App.entry_pdf.configure(state="normal",text_color=WHITE)
        self.App.button_browse_pdf.configure(state="normal")
        self.App.checkbox_extract.configure(state="normal")
        self.checkbox_extract_var.set(self.App.CHECK_EXTRACT)
    
    def glob_enable_extract(self):
        self.enable_extract()
        self.disable_create()
        self.disable_update()
        self.App.update()
    def glob_enable_create(self):
        self.enable_create()
        self.disable_update()
        self.disable_extract()
        self.App.update()
    def glob_enable_update(self):
        self.enable_update()
        self.disable_extract()
        self.disable_create()
        self.App.update()

    def disable_excel(self):
        self.App.label_excel_dir.configure(text_color=GREY)
        self.App.entry_excel_dir.configure(state="disabled",text_color=GREY)
        self.App.button_browse_excel_dir.configure(state="disabled")
    def enable_excel(self):
        self.App.label_excel_dir.configure(text_color=WHITE)
        self.App.entry_excel_dir.configure(state="normal",text_color=WHITE)
        self.App.button_browse_excel_dir.configure(state="normal")
        
    def init_gif(self):
       # Load the original image
        self.loading_img = Image.open(os.path.join(self.asset_path,'loading.png'))
        self.loading_ctkimg = ctk.CTkImage(light_image=self.loading_img, size=[45,45])
         # Create a label to display it
        self.loading_label = ctk.CTkLabel(self.App, image=self.loading_ctkimg, text="")
        self.loading_angle = 0

    def cancel(self):
        self.App.destroy()
    
    

    
    ########## COMMAND ##########
    def open_settings(self):
        print("Settings selected")

    def open_help(self):
        print("Help selected")

    def browse_pdf(self):
        file_paths = filedialog.askopenfilenames(title="Sélectionner un fichier PDF", filetypes=[("PDF Files", "*.pdf")])
        if file_paths:
            self.pdf_path.set("; ".join(file_paths))
    
    def browse_excel_dir(self):
        dir_path = filedialog.askdirectory(title="Sélectionner un dossier")
        if dir_path:
            self.excel_dir_path.set(dir_path)
            
    def get_wb_from_year_var(self):
        year = self.year_var.get()
        excel_file_path = os.path.join(self.excel_dir_path.get(), f'{year}.xlsx')
        return load_workbook(filename=excel_file_path)
        
    def get_existing_sheet_from_year(self):
        try:
            wb = self.get_wb_from_year_var()
            sheet_names = [item for item in wb.sheetnames if item!=TOOL_SHEET_NAME]
            return sheet_names
        except Exception:
            return ['']
        
    def free_year_wb(self):
        try:
            wb = self.get_wb_from_year_var()
            threading.Thread(target=wb.close).start()
            messagebox.showinfo('Info','Workbook closed')
        except Exception:
            pass
        
    def update_combo_box(self, *args):
        sheet_names = self.get_existing_sheet_from_year()
        self.App.entry_month.configure(values=sheet_names)
        
    def on_checkbox_toggle(self, checkbox_var):
    # Check if the event was triggered by the "Extract" checkbox
        if checkbox_var == self.checkbox_extract_var:
                self.glob_enable_extract()
    # Check if the event was triggered by the "Create" checkbox
        elif checkbox_var == self.checkbox_create_var:
                self.glob_enable_create()
    # Check if the event was triggered by the "Update" checkbox
        elif checkbox_var == self.checkbox_update_var:
                self.glob_enable_update()
                self.update_combo_box()
                

    def background_task(self):
        """Simulate background work and stop loading after 5s."""
        self.running_thread = True
        self.is_processing = True
        time.sleep(5)
        self.is_processing = False
        self.running_thread = False
        
    def ok(self):
        if (self.checkbox_update_var.get() == 0 and self.checkbox_create_var.get() == 0 and
                                    self.checkbox_extract_var.get() == self.App.CHECK_EXTRACT): # Extraction (Create)
            threading.Thread(target=self.launch_creation_extraction).start()
            self.running_thread = True
            self.start_loading()
        elif (self.checkbox_update_var.get() == 0 and self.checkbox_create_var.get() == self.App.CHECK_CREATE and
                                    self.checkbox_extract_var.get() == 0): # Create
            threading.Thread(target=self.launch_creation).start()
            self.running_thread = True
            self.start_loading()
        elif (self.checkbox_update_var.get() == self.App.CHECK_UPDATE and self.checkbox_create_var.get() == 0 and
                                    self.checkbox_extract_var.get() == 0): # Update/Realisation
            threading.Thread(target=self.launch_realisation).start()
            self.running_thread = True
            self.start_loading()
            
    def launch_creation(self):
        self.running_thread = True
        self.is_processing = True
        self.year = normalize_string(self.year_create.get())
        self.month = normalize_string(self.month_create.get())
        self.manage_excel_file("Extraction")
        if self.wb and self.ws:
            if not self.managment.ws_creat:
                self.is_processing = False
                messagebox.showinfo("Info", f"This sheet ({self.managment.ws_name}) already exist in {self.managment.file_path}")
            else:
                mks = SheetMarker(self.ws)
                set_all_data_validation(self.wb,mks)
                is_save_ok = self.managment.save_wb()
                self.is_processing = False
                messagebox.showinfo("Success", f"Sheet ({self.managment.ws_name}) created here: {self.managment.file_path}") if is_save_ok else None
        else:
            self.is_processing = False
            messagebox.showerror("Error", f"ERROR during Creation for {self.managment.ws_name}!")
        self.wb.close()
        self.running_thread = False
        
    def launch_creation_extraction(self):
        self.running_thread = True
        file_paths_list = [path.strip() for path in self.pdf_path.get().split(';')]
        for path_pdf in file_paths_list:
            self.is_processing = True
            extraction = PDFExtraction(path_pdf)
            self.year = normalize_string(extraction.year)
            self.month = normalize_string(extraction.month)
            self.manage_excel_file("Extraction")
            f_update = True
            if self.wb and self.ws:
                if not self.managment.ws_creat:
                    self.is_processing = False
                    if not messagebox.askyesno("Confirmation", f"This sheet ({self.managment.ws_name}) already exist in {self.managment.file_path}, do you want to update it ?"):
                        f_update = False
                if f_update:
                    self.is_processing = True
                    StyleCell(self.wb)
                    mks = SheetMarker(self.ws)
                    WriteExtraction(self.ws,extraction,mks)
                    Valid(self.ws,mks)
                    set_all_data_validation(self.wb,mks)
                    is_save_ok = self.managment.save_wb()
                    self.is_processing = False
                    messagebox.showinfo("Success", f"Extraction for {self.managment.ws_name} completed successfully!") if is_save_ok else None
            else:
                self.is_processing = False
                messagebox.showerror("Error", f"ERROR during Extraction for {self.managment.ws_name}!")
        self.wb.close()
        run_excel_path(self.file_path)
        self.running_thread = False
        return self.ws

    def launch_realisation(self):
        self.running_thread = True
        self.is_processing = True
        self.year = normalize_string(self.year_var.get())
        self.month = normalize_string(self.month_var.get().lower())
        self.manage_excel_file("Realisation")
        if self.wb:
            if self.ws:
                self.is_processing = False
                if messagebox.askyesno("Confirmation", f"{self.managment.ws_name} sheet already exist here:\n{self.managment.file_path}\n\nDo you want to update it ?"):
                    self.is_processing = True
                    StyleCell(self.wb)
                    mks = SheetMarker(self.ws)
                    realisation_month = MonthRealisation(self.ws,mks,"Monthly")
                    realisation_pdf = MonthRealisation(self.ws,mks,"Pdf")
                    realisation_intern = MonthRealisation(self.ws,mks,"Intern")
                    WriteRealisation(self.ws,mks,realisation_month,realisation_pdf,realisation_intern)
                    Valid(self.ws,mks)
                    set_all_data_validation(self.wb,mks)
                    is_save_ok = self.managment.save_wb()
                    self.is_processing = False
                    messagebox.showinfo("Success", f"Operation completed successfully for {self.managment.ws_name}!") if is_save_ok else None
                    self.wb.close()
                    run_excel_path(self.file_path)
            else:
                self.is_processing = False
                messagebox.showinfo("Info", f"Create the sheet {self.month} before")
        else:
            self.is_processing = False
            messagebox.showinfo("Info", f"Build the workbook {self.year} before")
        self.running_thread = False
        
    def manage_excel_file(self,man_type):
        self.managment = ManageExcelFile(self.excel_dir_path.get(),self.year,self.month,man_type)
        self.wb,self.ws,self.file_path = self.managment.get_wb_ws_fpath()

    def start_loading(self):
        self.disable_excel()
        self.disable_update()
        self.disable_extract()
        self.disable_create()
        self.animate_gif()
            
    def animate_gif(self):
        """Animate image rotation while not paused."""
        if self.is_processing:
            # Si l'image n'est pas déjà packée, la packer
            if not self.loading_label.winfo_ismapped():
                self.loading_label.place(x=15, y=self.App.winfo_height()-15, anchor="sw")

            # Rotation de l'image et mise à jour
            rotated = self.loading_img.rotate(self.loading_angle, resample=Image.BICUBIC)
            self.loading_ctkimg = ctk.CTkImage(light_image=rotated, size=(45, 45))
            self.loading_label.configure(image=self.loading_ctkimg)

            # Mise à jour de l'angle
            self.loading_angle = (self.loading_angle + 10) % 360
        else:
            # Si l'image est packée et is_processing est False, on oublie le pack
            if self.loading_label.winfo_ismapped():
                self.loading_label.place_forget()

        if self.running_thread:
            # Re-planifier la mise à jour toutes les 100 ms
            self.App.after(100, self.start_loading)
        else:
             # Autres actions à faire une fois que la tâche est terminée
            self.enable_excel()
            if self.checkbox_update_var.get() == self.App.CHECK_UPDATE:
                self.glob_enable_update()
            elif self.checkbox_extract_var.get() == self.App.CHECK_EXTRACT:
                self.glob_enable_extract()



if __name__ == "__main__":
    app = Main()
    app.App.mainloop()
