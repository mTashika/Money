import unicodedata
from openpyxl.utils import range_boundaries,get_column_letter
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
import Const as C
import Const_txt_excel as CT
from Const_Balise_Excel import EXT_VAL
import re
from openpyxl.comments import Comment
import subprocess

def month_number_to_name(month_number):
    month_number = int(month_number)
    months = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
        5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
        9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
    }
    return months.get(month_number, "Mois invalide")

def set_all_data_validation(wb,mkrs):
    for ws in wb.worksheets:
        if ws.title != C.TOOL_SHEET_NAME:
            dv1 = DataValidation(type="list", formula1=f"={C.TOOL_SHEET_NAME}!$A:$A")
            dv2 = DataValidation(type="list", formula1=f"={C.TOOL_SHEET_NAME}!$B:$B")
            dv3 = DataValidation(type="list", formula1='"' + ','.join(C.VALID_TYPE) + '"',
                    showErrorMessage=True,
                    errorTitle = CT.DATA_VAL_ERROR_INTERN_TITLE,
                    error = CT.DATA_VAL_ERROR_INTERN_MSG)
            dv4 = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", showDropDown=False,
                    showErrorMessage=True,
                    errorTitle = CT.DATA_VAL_ERROR_SHARE_TITLE,
                    error = CT.DATA_VAL_ERROR_SHARE_MSG)
            
            ws.add_data_validation(dv1)
            ws.add_data_validation(dv2)
            ws.add_data_validation(dv3)
            ws.add_data_validation(dv4)

            for row in range(mkrs.B_DETAIL_LINE_ST, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value == EXT_VAL:
                    break
                dv1.add(ws.cell(row=row,column = mkrs.B_ID_C1_COL))
                dv2.add(ws.cell(row=row,column = mkrs.B_ID_C2_COL))
                dv3.add(ws.cell(row=row,column = mkrs.B_ID_TYPE_COL))
                dv4.add(ws.cell(row=row,column = mkrs.B_ID_REFUND_COL))

def check_cell_value(cell_value):
    if cell_value is None:
        return -1
    if isinstance(cell_value, str):
        return 0
    if isinstance(cell_value, int):
        return 1
    if isinstance(cell_value, float):
        return 2

def clear_zone(ws,minrow,maxrow,mincol,maxcol):        
        mcr_coord_list = [mcr.coord for mcr in ws.merged_cells.ranges]
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            if min_col>=mincol and max_col<=maxcol and min_row>=minrow and max_row<=maxrow:
                ws.unmerge_cells(mcr)
        for row in range(minrow, maxrow + 1):
            for col in range(mincol, maxcol + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.style = 'Normal'
                cell.comment = None

def unmerge_cells_by_coords(ws, start_row, start_col, end_row, end_col):
    start_cell = get_column_letter(start_col) + str(start_row)
    end_cell = get_column_letter(end_col) + str(end_row)
    ws.unmerge_cells(f'{start_cell}:{end_cell}')

def generate_table(t_size, val_max):
    table = []
    for i in range(t_size):
        table.append((i % val_max)+1)  # Génère un nombre de 1 à x en boucle
    return table
def remove_accent(input_str):
    return ''.join(
        c for c in unicodedata.normalize('NFD', input_str)
        if unicodedata.category(c) != 'Mn'
    )
    
def normalize_string(input_str):
    # Convert to lowercase
    lower_str = input_str.lower()
    # Remove accents
    normalized_str = remove_accent(lower_str)
    return normalized_str

def center_window(win, width, height):
        # Obtenir la taille de l'écran
        screen_width = win.winfo_screenwidth()
        screen_height = win.winfo_screenheight()

        # Calculer les coordonnées x et y pour centrer la fenêtre
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

        # Définir la taille et la position de la fenêtre
        win.geometry(f"{width}x{height}+{x}+{y}")
        
def get_current_month():
    return datetime.now().strftime("%B")
def get_current_year():
    return datetime.now().strftime("%Y")

def evaluate_excel_formula(formula: str, ws):
    """
    Evaluate a simple Excel formula using values from the worksheet.
    Only supports basic arithmetic operations (+, -, *, /).
    """
    if not formula.startswith("="):
        try:
            return float(formula)
        except ValueError:
            return formula  # Non-numeric value, return as-is

    # Remove leading '=' and extract cell references (e.g., A1, B2)
    expression = formula[1:]
    cell_refs = re.findall(r'[A-Z]+\d+', expression)

    # Replace each cell reference with its actual value from the worksheet
    for ref in cell_refs:
        cell_value = ws[ref].value
        if cell_value is None:
            cell_value = 0
        elif isinstance(cell_value, str) and cell_value.startswith("="):
            # Handle nested formulas recursively
            cell_value = evaluate_excel_formula(cell_value, ws)
        expression = expression.replace(ref, str(cell_value))

    try:
        # Evaluate the final expression safely
        return eval(expression)
    except Exception as e:
        raise ValueError(f"Invalid formula or error during evaluation: {e}")

def add_excel_comment(cell, text: str, author: str = "Auto"):
    """
    Add a comment (note) to a given cell in the worksheet.
    
    Parameters:
        cell cell in the worksheet
        text (str): The content of the comment.
        author (str): The author of the comment.
    """
    comment = Comment(text, author)
    cell.comment = comment


def run_excel_path(file_path):
    subprocess.run(['start', '', file_path], shell=True)

class FinancialOperation:
    """ Classe used for the data extraction from a PDF. """
    def __init__(self,num,type,name,value,unit='€'):
        self.num = num
        self.type = type
        self.name=name
        self.value = value
        self.unit = unit
    def print(self):
        print(f'{self.num}\t{self.name}: {self.value}{self.unit} ({self.type})')
