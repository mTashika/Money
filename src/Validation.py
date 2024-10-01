from Tools import check_cell_value
from Const import EXT_BILAN_TXT_VALID,VALD_TOT_POUR_ERROR,VALD_TOT_SEUIL_ERROR
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill,Font
from openpyxl.formatting.rule import CellIsRule

class Valid():
    def __init__(self,ws,mks):
        val_extraction_value = ws.cell(row=mks.b_ext_verif[0],column=mks.b_ext_verif[1]).value
        sold_re = ws.cell(row=mks.B_BILAN_ED_SOLDRE[0],column=mks.B_BILAN_ED_SOLDRE[1]).value
        cell_val_tot = ws.cell(row=mks.B_BILAN_ED_VAL[0],column=mks.B_BILAN_ED_VAL[1])
        
        if (not val_extraction_value or check_cell_value(sold_re) not in [1,2]):
            cell_val_tot.style = 'Neutral'
        else:
            cell_soldex = get_column_letter(mks.B_BILAN_ED_SOLDEX[1]) + str(mks.B_BILAN_ED_SOLDEX[0])
            cell_soldre = get_column_letter(mks.B_BILAN_ED_SOLDRE[1]) + str(mks.B_BILAN_ED_SOLDRE[0])
            abs_difference = f'ABS({cell_soldex} - {cell_soldre})'
            max_value = f'MAX(ABS({cell_soldex}), ABS({cell_soldre})) * {VALD_TOT_POUR_ERROR}'

            cell_val_tot.value = f'={abs_difference} <= MAX({VALD_TOT_SEUIL_ERROR}, {max_value})'
            
        cell_val_tot.number_format  = f';;;"{EXT_BILAN_TXT_VALID}"'
        cell_val_tot.alignment=Alignment(horizontal='center', vertical='center')
        rule_good = CellIsRule(operator='equal', formula=[True], stopIfTrue=True, fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),font = Font(color="006100"))
        rule_bad = CellIsRule(operator='equal', formula=[False], stopIfTrue=True, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),font = Font(color="9C0006"))
        ws.conditional_formatting.add(f'{get_column_letter(mks.B_BILAN_ED_VAL[1])}{mks.B_BILAN_ED_VAL[0]}', rule_good)
        ws.conditional_formatting.add(f'{get_column_letter(mks.B_BILAN_ED_VAL[1])}{mks.B_BILAN_ED_VAL[0]}', rule_bad)