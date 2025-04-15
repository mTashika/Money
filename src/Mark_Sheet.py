import Const_Balise_Excel as BE
from Tools import check_cell_value
class SheetMarker():
    def __init__(self,sheet):
        self.sheet = sheet
        self.B_BILAN_DATE = None
        self.B_BILAN_ST_SOLD = None
        self.B_BILAN_ED_SOLDEX = None
        self.B_BILAN_ED_SOLDRE = None
        self.B_BILAN_ED_VAL = None

        self.B_LOS_IN_ST_LINE = None # income/loss start line      
        self.B_LOS_IN_ED_LINE_CLEAR = None # income/loss end line

        self.B_IN_NAME_COL = None # income name column
        self.B_IN_EX_COL = None # income expected column
        self.B_IN_RE_COL = None # income real column
        self.B_IN_TOT_EX = None # income total expected cell
        self.B_IN_TOT_REAL = None # income total real cell
        self.B_in_ed_line = None # income end line
        
        self.B_LOS_NAME_COL = None # loss name column
        self.B_LOS_EX_COL = None # loss expected column
        self.B_LOS_RE_COL = None # loss real column
        self.B_LOS_TOT_EX = None # loss total expected cell
        self.B_LOS_TOT_REAL = None # loss total real cell
        self.B_los_ed_line = None # loss end line
        
        self.B_DETAIL_LINE_ST = None # detail line start
        self.B_ID_C1_COL = None # identification categorie 1 column
        self.B_ID_C2_COL = None # identification categorie 2 column
        self.B_ID_TYPE_COL = None # identification Type column
        self.B_ID_REFUND_COL = None # identification Refund column
        
        self.B_EXT_NAME_COL = None # extraction name column
        self.B_EXT_VALUE_COL = None # extraction value column
        self.B_ext_ed_line = None # extraction end line
        self.b_ext_sold_ed_est = None # extraction sold cell
        self.b_ext_sold_ed_true = None # extraction sold end true cell
        self.b_ext_verif = None # extraction verif cell
        
        self.B_REAL_COL_ST = None # realisation column start
        self.B_REAL_COL_ED = None # realisation column end
        
        self.B_VALID_EXT_PATH = None # validity extraction path cell
        self.B_VALID_EXT_DATE = None # validity extraction date cell
        self.B_VALID_ID_DATE = None # validity identification date cell
        self.B_VALID_REAL_DATE = None # validity realisation date cell
        
        self.extract_balise()
        a=1
        
    def extract_balise(self):
        for row in range(1, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if cell_value == BE.BILAN_DATE:
                    self.B_BILAN_DATE = [row,col+1]
                elif cell_value == BE.BILAN_ST_SOLD:
                    self.B_BILAN_ST_SOLD = [row,col+1]
                elif cell_value == BE.BILAN_ED_SOLD:
                    self.B_BILAN_ED_SOLDEX = [row+1,col+1]
                    self.B_BILAN_ED_SOLDRE = [row+1,col+2]
                    self.B_BILAN_ED_VAL = [row-1,col+3]
                elif cell_value == BE.INCOME:
                    self.B_IN_NAME_COL = col
                    self.B_IN_EX_COL = col+1
                    self.B_IN_RE_COL = col+2
                    self.B_LOS_IN_ST_LINE = row+4
                    self.B_IN_TOT_EX = [row+1,col+1]
                    self.B_IN_TOT_REAL = [row+1,col+2]
                elif cell_value == BE.LOSS:
                    self.B_LOS_NAME_COL = col
                    self.B_LOS_EX_COL = col+1
                    self.B_LOS_RE_COL = col+3
                    self.B_LOS_IN_ST_LINE = row+4
                    self.B_LOS_TOT_EX = [row+1,col+1]
                    self.B_LOS_TOT_REAL = [row+1,col+3]
                elif cell_value == BE.DETAIL:
                    self.B_LOS_IN_ED_LINE_CLEAR = row-1
                    self.B_DETAIL_LINE_ST = row+2
                    self.B_ID_C1_COL = col
                    self.B_ID_C2_COL = col+1
                    self.B_ID_TYPE_COL = col+2
                    self.B_ID_REFUND_COL = col+3
                    self.B_EXT_NAME_COL = col+4
                    self.B_EXT_VALUE_COL = col+5
                    self.B_REAL_COL_ST = col+7
                    self.B_REAL_COL_ED = self.B_REAL_COL_ST+8
                elif cell_value == BE.VALIDITY:
                    self.B_VALID_EXT_PATH = [row+1,col+2]
                    self.B_VALID_EXT_DATE = [row+2,col+2]
                    self.B_VALID_ID_DATE = [row+3,col+2]
                    self.B_VALID_REAL_DATE = [row+4,col+2]
                elif cell_value == BE.EXT_VAL:
                    self.b_ext_sold_ed_est = [row,col+1]
                    self.b_ext_verif = [row,col+2]

        self.B_in_ed_line = self.B_LOS_IN_ED_LINE_CLEAR
        for row in range(self.B_LOS_IN_ST_LINE, self.B_DETAIL_LINE_ST):
            cv = self.sheet.cell(row=row, column=self.B_IN_NAME_COL).value
            if check_cell_value(cv) != 0: # None, int or float
                self.B_in_ed_line = row-1
                if self.B_in_ed_line<self.B_LOS_IN_ST_LINE:
                    self.B_in_ed_line=self.B_LOS_IN_ST_LINE
                break

        self.B_los_ed_line = self.B_LOS_IN_ED_LINE_CLEAR    
        for row in range(self.B_LOS_IN_ST_LINE, self.B_DETAIL_LINE_ST):
            cv = self.sheet.cell(row=row, column=self.B_LOS_NAME_COL).value
            if check_cell_value(cv) != 0: # None, int or float
                self.B_los_ed_line = row-1
                if self.B_los_ed_line < self.B_LOS_IN_ST_LINE:
                    self.B_los_ed_line = self.B_LOS_IN_ST_LINE
                break
            
        if self.b_ext_sold_ed_est is not None:
            for row in range(self.B_DETAIL_LINE_ST, self.b_ext_sold_ed_est[0]):
                cv = self.sheet.cell(row=row, column=self.B_EXT_NAME_COL).value
                if check_cell_value(cv) != 0:
                    self.B_ext_ed_line = row-1
                    if self.B_ext_ed_line<self.B_DETAIL_LINE_ST:
                        self.B_ext_ed_line=self.B_DETAIL_LINE_ST

            
    def print(self):
        for attr, value in vars(self).items():
            if value is not None:
                print(f"{attr}: {value}")
            else:
                print(f"{attr}: None")

if __name__ == "__main__":
    import openpyxl
    path = r"C:\Users\mcast\OneDrive\Bureau\Classeur1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    s = SheetMarker(sheet)
    s.print()