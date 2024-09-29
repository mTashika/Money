
from Const import TOOL_SHEET_NAME,CAT_1,CAT_2,INIT_SHEET_CAT1_TITLE,INIT_SHEET_CAT2_TITLE

class InitWorkbook():
    def __init__(self,wb):
        self.wb = wb
        self.init_tool_sheet()
        
    def init_tool_sheet(self):
        # protect all the sheet
        ws_tools = self.wb.create_sheet(title=TOOL_SHEET_NAME)
        ws_tools.cell(1,1).value = INIT_SHEET_CAT1_TITLE
        ws_tools.cell(1,2).value = INIT_SHEET_CAT2_TITLE
        for idx, string in enumerate(CAT_1, start=2):
            ws_tools.cell(row=idx, column=1).value = string
        for idx, string in enumerate(CAT_2, start=2):
            ws_tools.cell(row=idx, column=2).value = string
        
        
if __name__ == "__main__":
    import openpyxl
    from Styles import StyleCell
    
    path = r"D:\One Drive\OneDrive\Bureau\Classeur1.xlsx"
    workbook = openpyxl.load_workbook(path)
    StyleCell(workbook)
    InitWorkbook(workbook)
    workbook.save(path)