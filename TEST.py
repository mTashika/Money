from openpyxl import Workbook

from openpyxl.styles.protection import Protection

def protect_worksheet(ws):
    """
    Protect the worksheet so only cells in column A are editable.
    Users can select and edit unlocked cells only, but cannot format or move anything.
    """
    # Unlock cells in column A (e.g., A1 to A100)
    for row in ws.iter_rows(min_col=1, max_col=1, max_row=100):
        for cell in row:
            cell.protection = Protection(locked=False)

    # Apply protection with limited permissions
    ws.protection.sheet = True
    ws.protection.password = "123"

    # Set permission options
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False

    # # Allow selection of unlocked cells ONLY
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False




wb = Workbook()
ws = wb.active

# Add test data
for i in range(1, 6):
    ws[f"A{i}"] = f"Editable {i}"
    ws[f"B{i}"] = f"Protected {i}"

# Apply protection
protect_worksheet(ws)

# Save workbook
wb.save(r"C:\Users\mcast\OneDrive\Bureau\protected_with_function.xlsx")
