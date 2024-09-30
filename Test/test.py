from openpyxl import Workbook
from openpyxl.worksheet.protection import SheetProtection

def protect_ws(ws):
    # Activer la protection de la feuille
    ws.protection = SheetProtection(sheet=True, formatCells=True)
    
    # Empêcher la sélection des cellules verrouillées (ne permettre que la sélection des cellules non protégées)
    ws.protection.enableSelection = 'unlocked'  # Empêcher la sélection des cellules verrouillées

# Exemple d'utilisation
wb = Workbook()
ws = wb.active

# Exemple de données
ws['A1'] = "Donnée protégée"
ws['B1'] = "Donnée non protégée"

# Déverrouiller certaines cellules (par exemple, la cellule B1)
ws['B1'].protection = ws['B1'].protection.copy(locked=False)

# Appeler la fonction pour protéger la feuille
protect_ws(ws)

# Enregistrer le fichier
wb.save('feuille_protegee_selection_limitee.xlsx')
